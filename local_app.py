#!/usr/bin/env python3
"""Local web app for generating timeline Excel files."""

from __future__ import annotations

import argparse
import base64
import json
import os
import re
import tempfile
import webbrowser
import xml.etree.ElementTree as ElementTree
from datetime import date, datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter

from scripts.build_timeline import build_workbook


HOST = "127.0.0.1"
DEFAULT_PORT = 8765


def normalize_base_path(value: str) -> str:
    base = value.strip().rstrip("/")
    if base and not base.startswith("/"):
        base = "/" + base
    return base


BASE_PATH = normalize_base_path(os.environ.get("BASE_PATH", ""))


def with_base(path: str) -> str:
    if not path.startswith("/"):
        path = "/" + path
    if not BASE_PATH:
        return path
    return BASE_PATH + path


def strip_base_path(path: str) -> str:
    if BASE_PATH and path.startswith(BASE_PATH + "/"):
        stripped = path[len(BASE_PATH) :]
        return stripped or "/"
    return path


def render_index_html() -> str:
    return (
        INDEX_HTML
        .replace("__BASE_PATH_JSON__", json.dumps(BASE_PATH))
        .replace("__LOGO_SRC__", with_base("/assets/kivisense-logo.png"))
    )


INDEX_HTML = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>项目排期工作台</title>
  <style>
    :root {
      --bg: #f5f7fb;
      --card: #ffffff;
      --card-soft: #f8fbff;
      --text: #182033;
      --muted: #667085;
      --muted-2: #98a2b3;
      --line: #dfe6f1;
      --line-2: #edf1f7;
      --primary: #0f62df;
      --primary-dark: #084fb8;
      --primary-soft: #eef5ff;
      --green: #079455;
      --green-soft: #e8f8ef;
      --blue: #1b64da;
      --blue-soft: #eaf2ff;
      --yellow: #b77900;
      --yellow-soft: #fff3d6;
      --orange: #d9480f;
      --orange-soft: #fff0e8;
      --gray: #475467;
      --gray-soft: #f2f4f7;
      --danger: #d92d20;
      --danger-soft: #fff1f0;
      --shadow: 0 18px 48px rgba(15, 23, 42, 0.08);
      --radius: 18px;
    }
    * { box-sizing: border-box; }
    html, body { min-height: 100%; }
    body {
      margin: 0;
      font-family: Inter, "Google Sans Text", -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft YaHei", sans-serif;
      color: var(--text);
      background:
        radial-gradient(circle at 20% 0%, rgba(15, 98, 223, 0.08), transparent 34%),
        linear-gradient(180deg, #fbfcff 0%, var(--bg) 52%, #f1f4f9 100%);
      padding-bottom: 88px;
      overflow-x: hidden;
    }
    button, input, select { font: inherit; }
    button { cursor: pointer; }
    .app { width: min(1560px, 100%); margin: 0 auto; padding: 0 24px 28px; }
    .topbar {
      min-height: 72px;
      display: flex; align-items: center; justify-content: space-between; gap: 18px;
      padding: 15px 0 16px;
    }
    .brand { display: flex; align-items: center; gap: 14px; min-width: 250px; }
    .brand-mark {
      width: 42px; height: 42px; border-radius: 14px;
      display: grid; place-items: center;
      color: #fff; font-weight: 900;
      background: linear-gradient(135deg, #4776f5, #1457d9);
      box-shadow: 0 10px 24px rgba(15, 98, 223, 0.24);
      flex: 0 0 auto;
    }
    .brand-title { margin: 0; font-size: 24px; line-height: 1.1; letter-spacing: -0.04em; }
    .brand-subtitle { margin: 5px 0 0; color: var(--muted); font-size: 13px; }
    .header-actions { display: flex; align-items: center; justify-content: flex-end; gap: 12px; min-width: 0; flex-wrap: wrap; }
    .project-title-box { display: flex; align-items: center; gap: 8px; min-width: min(560px, 48vw); }
    .project-title-box label { color: var(--muted); font-size: 12px; font-weight: 900; white-space: nowrap; }
    .project-title-box input { width: min(420px, 36vw); min-width: 260px; height: 40px; }
    input, select {
      border: 1px solid var(--line);
      background: #fff;
      color: var(--text);
      outline: none;
      border-radius: 12px;
      padding: 0 12px;
      transition: border-color .16s ease, box-shadow .16s ease, background .16s ease;
    }
    input::placeholder { color: #9aa7bb; }
    input:focus, select:focus {
      border-color: rgba(15, 98, 223, 0.55);
      box-shadow: 0 0 0 4px rgba(15, 98, 223, 0.10);
    }
    .segmented {
      display: inline-flex; align-items: center; gap: 2px; padding: 4px; border: 1px solid var(--line);
      background: rgba(255,255,255,.78); border-radius: 999px; box-shadow: 0 8px 24px rgba(15, 23, 42, 0.06);
      flex: 0 0 auto;
    }
    .segmented button { min-width: 72px; height: 32px; padding: 0 14px; border: 0; border-radius: 999px; color: var(--muted); background: transparent; font-weight: 850; white-space: nowrap; }
    .segmented button.active { color: #fff; background: var(--primary); box-shadow: 0 8px 18px rgba(15, 98, 223, .25); }
    .btn {
      min-width: 96px; height: 40px; border-radius: 13px; border: 1px solid var(--line); background: #fff; color: var(--text);
      padding: 0 16px; font-weight: 850; display: inline-flex; align-items: center; justify-content: center; gap: 8px;
      transition: transform .15s ease, box-shadow .15s ease, background .15s ease, border-color .15s ease;
      white-space: nowrap;
    }
    .btn:hover { transform: translateY(-1px); border-color: #cfd7e6; box-shadow: 0 10px 22px rgba(15, 23, 42, .08); }
    .btn.primary { background: var(--primary); color: #fff; border-color: var(--primary); box-shadow: 0 12px 26px rgba(15, 98, 223, .22); }
    .btn.primary:hover { background: var(--primary-dark); }
    .btn.small { min-width: 0; height: 34px; padding: 0 12px; border-radius: 11px; font-size: 12px; }
    .layout { display: grid; grid-template-columns: minmax(0, 1fr) 248px; gap: 14px; align-items: start; }
    .panel { background: rgba(255,255,255,.94); border: 1px solid var(--line); border-radius: var(--radius); box-shadow: var(--shadow); overflow: hidden; }
    .panel-head { display: flex; align-items: flex-start; justify-content: space-between; gap: 12px; padding: 18px 22px 16px; border-bottom: 1px solid var(--line-2); }
    .panel-title { margin: 0; font-size: 18px; line-height: 1.2; letter-spacing: -0.02em; }
    .panel-desc { margin: 6px 0 0; color: var(--muted); font-size: 13px; line-height: 1.45; }
    .task-panel .panel-desc { display: none; }
    .panel-tools { display: flex; align-items: center; gap: 8px; flex-wrap: wrap; justify-content: flex-end; }
    .message { margin: 16px 22px 0; padding: 11px 14px; border-radius: 12px; font-size: 13px; font-weight: 800; display: none; }
    .message.show { display: block; }
    .message.success { color: #05603a; border: 1px solid #8be0b0; background: #edfcf3; }
    .message.error { color: var(--danger); border: 1px solid #fda29b; background: var(--danger-soft); }
    .stage-workbench {
      display: grid;
      grid-template-columns: minmax(0, 1fr) minmax(240px, 320px);
      gap: 14px;
      align-items: center;
      padding: 14px 22px;
      border-bottom: 1px solid var(--line-2);
      background: linear-gradient(180deg, #ffffff, #fbfdff);
    }
    .stage-tabs {
      display: flex;
      align-items: center;
      gap: 8px;
      min-width: 0;
      overflow-x: auto;
      padding: 2px 0 4px;
      scrollbar-width: thin;
    }
    .stage-tab {
      flex: 0 0 auto;
      min-width: 92px;
      height: 36px;
      padding: 0 14px;
      border: 1px solid var(--line);
      border-radius: 999px;
      background: #fff;
      color: #475569;
      font-size: 13px;
      font-weight: 850;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 8px;
      white-space: nowrap;
      transition: background .15s ease, border-color .15s ease, color .15s ease, box-shadow .15s ease;
    }
    .stage-tab:hover { border-color: #b7ccff; background: var(--primary-soft); color: var(--primary); }
    .stage-tab.active { background: #111827; border-color: #111827; color: #fff; box-shadow: 0 10px 22px rgba(15, 23, 42, .16); }
    .stage-tab[draggable="true"] { cursor: grab; user-select: none; }
    .stage-tab[draggable="true"]:active { cursor: grabbing; }
    .stage-tab.stage-dragging { opacity: .5; }
    .stage-tab.stage-drag-over-before { box-shadow: inset 3px 0 0 var(--primary); }
    .stage-tab.stage-drag-over-after { box-shadow: inset -3px 0 0 var(--primary); }
    .stage-tab-count {
      min-width: 22px;
      height: 22px;
      padding: 0 7px;
      border-radius: 999px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      background: rgba(148,163,184,.16);
      color: inherit;
      font-size: 11px;
      font-weight: 900;
    }
    .stage-tab.active .stage-tab-count { background: rgba(255,255,255,.18); }
    .stage-name-control {
      min-width: 0;
      display: grid;
      grid-template-columns: auto minmax(0, 1fr);
      align-items: center;
      gap: 10px;
      justify-self: end;
      width: 100%;
    }
    .stage-name-control span {
      color: var(--muted);
      font-size: 12px;
      font-weight: 900;
      white-space: nowrap;
    }
    .stage-name-control input { width: 100%; height: 38px; border-radius: 12px; }
    .table-shell { padding: 0 0 14px; }
    .table-scroll { overflow-x: auto; overflow-y: visible; background: #fff; }
    table { width: 100%; min-width: 790px; border-collapse: separate; border-spacing: 0; table-layout: fixed; }
    th, td { height: 58px; padding: 10px 12px; border-bottom: 1px solid var(--line-2); vertical-align: middle; background: #fff; }
    th { height: 44px; color: #435168; background: #f8fafd; font-size: 11px; letter-spacing: .07em; text-transform: uppercase; font-weight: 900; text-align: left; position: sticky; top: 0; z-index: 2; }
    tbody tr:hover td { background: #fbfdff; }
    tbody tr.dragging td { opacity: .45; background: #f8fbff !important; }
    tbody tr.drag-over-before td { box-shadow: inset 0 3px 0 var(--primary); }
    tbody tr.drag-over-after td { box-shadow: inset 0 -3px 0 var(--primary); }
    tbody tr.group-start td { border-top: 10px solid #f4f7fb; }
    tbody tr.group-0 td { background: #ffffff; }
    tbody tr.group-1 td { background: #fbfdff; }
    tbody tr.group-2 td { background: #fffdf8; }
    tbody tr.group-3 td { background: #fbfffc; }
    tbody tr.group-4 td { background: #fffbfd; }
    tbody tr.group-5 td { background: #fcfbff; }
    tbody tr.group-1:hover td, tbody tr.group-2:hover td, tbody tr.group-3:hover td, tbody tr.group-4:hover td, tbody tr.group-5:hover td { background: #f7fbff; }
    .stage-pill { display:inline-flex; align-items:center; max-width:100%; min-height:32px; border:1px solid rgba(148,163,184,.24); border-radius:12px; padding:0 12px; font-weight:900; color:#253047; background: rgba(255,255,255,.72); }

    tbody tr:last-child td { border-bottom: 0; }
    td input, td select { width: 100%; height: 38px; border-radius: 12px; font-size: 13px; }
    .date-field {
      width: 100%; height: 38px; border-radius: 12px; border: 1px solid var(--line); background: #fff; color: var(--text);
      display: flex; align-items: center; justify-content: space-between; gap: 10px; padding: 0 12px;
      font-size: 13px; font-weight: 850; text-align: left; box-shadow: none; transition: border-color .16s ease, box-shadow .16s ease, background .16s ease;
    }
    .date-field:hover { border-color: #b9c7dd; background: #fbfdff; }
    .date-field:focus { outline: none; border-color: rgba(15, 98, 223, 0.58); box-shadow: 0 0 0 4px rgba(15, 98, 223, 0.11); }
    .date-field .date-placeholder { color: #98a2b3; font-weight: 850; }
    .date-field .date-icon { color: var(--primary); font-size: 15px; flex: 0 0 auto; }
    .calendar-popover {
      position: fixed; z-index: 100; width: 292px; padding: 14px; border: 1px solid #d9e4f5; border-radius: 18px;
      background: rgba(255,255,255,.98); box-shadow: 0 24px 60px rgba(15, 23, 42, .22); backdrop-filter: blur(14px);
    }
    .calendar-head { display:flex; align-items:center; justify-content:space-between; gap:10px; margin-bottom: 12px; }
    .calendar-title { font-size: 14px; font-weight: 950; color: #172033; letter-spacing: -0.01em; }
    .calendar-nav { display:flex; align-items:center; gap:6px; }
    .calendar-nav button, .calendar-day {
      border: 1px solid transparent; background: transparent; border-radius: 10px; color: #334155; font-weight: 850;
    }
    .calendar-nav button { width: 30px; height: 30px; color: var(--primary); background: var(--primary-soft); }
    .calendar-nav button:hover { border-color: #b7ccff; background:#e5efff; }
    .calendar-week, .calendar-grid { display:grid; grid-template-columns: repeat(7, 1fr); gap: 5px; }
    .calendar-week { margin-bottom: 6px; color:#8392a9; font-size:10px; font-weight:950; text-align:center; }
    .calendar-day { height: 32px; font-size: 12px; cursor:pointer; }
    .calendar-day:hover { background: var(--primary-soft); border-color:#b7ccff; color: var(--primary); }
    .calendar-day.muted { color:#bdc6d5; }
    .calendar-day.today { border-color:#9ebcff; }
    .calendar-day.selected { background: var(--primary); color:#fff; box-shadow:0 8px 18px rgba(15,98,223,.25); }
    .calendar-foot { display:flex; align-items:center; justify-content:space-between; gap:8px; margin-top: 12px; padding-top: 12px; border-top: 1px solid var(--line-2); }
    .calendar-foot button { height: 30px; border-radius: 10px; border: 1px solid var(--line); background:#fff; padding: 0 10px; font-size:12px; font-weight:900; color:#334155; }
    .calendar-foot button.primary-lite { background: var(--primary-soft); color: var(--primary); border-color:#b7ccff; }
    .col-stage { width: 112px; }
    .col-task { width: 310px; }
    .col-stakeholder { width: 160px; }
    .col-status { width: 132px; }
    .col-start, .col-end { width: 136px; }
    .col-days { width: 90px; }
    .col-actions { width: 170px; text-align: right; position: sticky; right: 0; z-index: 3; box-shadow: -12px 0 20px rgba(15, 23, 42, .05); padding-left: 8px; padding-right: 8px; }
    th.col-actions { background: #f8fafd; z-index: 4; }
    td.col-actions { background: #fff; }
    tbody tr:hover td.col-actions { background: #fbfdff; }
    .status-chip { display: inline-flex; align-items: center; justify-content: center; gap: 6px; min-width: 78px; height: 28px; border-radius: 999px; padding: 0 10px; font-size: 12px; font-weight: 900; }
    .status-chip::before { content: ''; width: 6px; height: 6px; border-radius: 50%; background: currentColor; }
    .status-chip.done { color: var(--green); background: var(--green-soft); }
    .status-chip.incomplete { color: var(--gray); background: var(--gray-soft); }
    .status-wrap { position: relative; display: inline-flex; }
    .status-wrap select { opacity: 0; position: absolute; inset: 0; cursor: pointer; }
    .days-pill { display: inline-flex; align-items: center; justify-content: center; min-width: 58px; height: 30px; border-radius: 999px; background: #f1f5f9; font-size: 12px; font-weight: 900; color: #26344a; }
    .days-pill.error { color: var(--danger); background: var(--danger-soft); min-width: 76px; }
    .row-actions { display: inline-flex; align-items: center; justify-content: flex-end; gap: 4px; flex-wrap: nowrap; width: 100%; }
    tr.dragging td { opacity: .58; background: #eef5ff !important; }
    .drag-handle { cursor: grab; color: #334155; }
    .drag-handle:active { cursor: grabbing; }
    .icon-btn { width: 26px; height: 28px; border-radius: 9px; border: 1px solid var(--line); background: #fff; color: #55657b; display: inline-grid; place-items: center; font-size: 12px; font-weight: 900; flex: 0 0 auto; }
    .icon-btn:hover { color: var(--primary); border-color: #b7ccff; background: var(--primary-soft); }
    .side-stack { display: grid; gap: 14px; }
    .side-card { padding: 13px; }
    .side-card h3 { margin: 0 0 6px; font-size: 14px; }
    .side-card p { margin: 0 0 10px; color: var(--muted); font-size: 11px; line-height: 1.45; }
    .template-list { display: grid; gap: 10px; }
    .template-card { border: 1px solid var(--line); border-radius: 13px; background: #fff; padding: 10px; text-align: left; transition: all .15s ease; width: 100%; }
    .template-card:hover, .template-card.active { border-color: #9ebcff; background: #f7faff; box-shadow: 0 10px 24px rgba(15, 98, 223, .08); }
    .template-card strong { display: block; font-size: 12px; margin-bottom: 4px; }
    .template-card span { display: block; color: var(--muted); font-size: 11px; line-height: 1.4; }
    .field-controls { display: grid; gap: 8px; }
    .check-row { height: 38px; display: flex; align-items: center; justify-content: space-between; gap: 12px; border: 1px solid var(--line); border-radius: 12px; padding: 0 12px; background: #fff; font-size: 13px; font-weight: 850; }
    .check-row input { width: 16px; height: 16px; accent-color: var(--primary); }
    .bottom-bar { position: fixed; left: 50%; bottom: 14px; transform: translateX(-50%); width: min(1560px, calc(100vw - 48px)); z-index: 30; display: flex; align-items: center; justify-content: space-between; gap: 14px; padding: 10px 14px; border: 1px solid rgba(219, 226, 236, .9); border-radius: 18px; background: rgba(255,255,255,.92); backdrop-filter: blur(16px); box-shadow: 0 16px 44px rgba(15, 23, 42, .10); }
    .quick-add { display: flex; align-items: center; gap: 8px; overflow-x: auto; min-width: 0; padding: 6px 8px; border: 1px solid var(--line-2); border-radius: 14px; background: #f8fbff; }
    .quick-add-label { flex: 0 0 auto; color: var(--muted); font-size: 12px; font-weight: 900; margin-right: 4px; }
    .quick-add .btn { flex: 0 0 auto; height: 32px; background: #fff; }
    .bar-actions { flex: 0 0 auto; display: flex; align-items: center; gap: 10px; }
    .import-year { width: 146px; min-width: 0; height: 38px; }
    .drawer-backdrop { position: fixed; inset: 0; background: rgba(15, 23, 42, .36); z-index: 50; opacity: 0; pointer-events: none; transition: opacity .18s ease; }
    .drawer-backdrop.open { opacity: 1; pointer-events: auto; }
    .preview-drawer { position: fixed; left: 50%; bottom: 0; transform: translate(-50%, 105%); width: min(1500px, calc(100vw - 32px)); height: min(62vh, 650px); z-index: 60; background: #fff; border: 1px solid #dbe5f3; border-bottom: 0; border-radius: 24px 24px 0 0; box-shadow: 0 -30px 80px rgba(15, 23, 42, .26); transition: transform .22s ease; overflow: hidden; display: flex; flex-direction: column; }
    .preview-drawer.open { transform: translate(-50%, 0); }
    .drawer-head {
      padding: 12px 18px;
      border-bottom: 1px solid var(--line-2);
      display: grid;
      grid-template-columns: auto minmax(220px, 1fr) auto 40px;
      align-items: center;
      gap: 14px;
      background: linear-gradient(180deg, rgba(248,251,255,.98), rgba(255,255,255,.98));
    }
    .drawer-title h2 { margin: 0; font-size: 18px; line-height: 1.15; letter-spacing: -0.02em; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
    .drawer-close { min-width: 38px; width: 38px; height: 38px; padding: 0; border-radius: 14px; }
    .drawer-stats { display: flex; align-items: center; gap: 8px; justify-content: flex-end; }
    .compact-stat { width: 96px; height: 42px; border: 1px solid #dce6f3; background: #fff; border-radius: 14px; padding: 6px 10px; box-shadow: 0 8px 18px rgba(15,23,42,.04); }
    .compact-stat .stat-label { display:block; color: var(--muted); font-size: 10px; line-height: 1; font-weight: 900; white-space: nowrap; }
    .compact-stat .stat-value { display:block; margin-top: 5px; font-size: 18px; line-height: 1; font-weight: 950; letter-spacing: -0.03em; }
    .drawer-tabs { display: inline-flex; gap: 6px; justify-content: flex-start; align-items: center; padding: 4px; border: 1px solid #dce6f3; border-radius: 999px; background: #f8fbff; }
    .tab { flex: 0 0 auto; height: 34px; min-width: 116px; border: 0; border-radius: 999px; background: transparent; color: var(--muted); font-weight: 900; }
    .tab.active { color: #fff; background: var(--primary); border-color: var(--primary); box-shadow: 0 10px 22px rgba(15,98,223,.22); }
    .drawer-body { min-height: 0; overflow: auto; padding: 16px 18px 24px; flex: 1; background: linear-gradient(180deg, #fbfdff, #ffffff); }
    .gantt-card { border: 1px solid #dbe5f3; border-radius: 20px; overflow: hidden; background: #fff; box-shadow: 0 18px 45px rgba(15,23,42,.06); }
    .timeline-scale { display: grid; grid-template-columns: 260px repeat(6, 1fr); gap: 0; color: #8392a9; font-size: 11px; font-weight: 950; padding: 14px 18px 8px; background: linear-gradient(180deg, #f8fbff, #fff); border-bottom: 1px solid #eef2f7; }
    .timeline-scale span:not(:first-child) { text-align: center; }
    .gantt-list { padding: 14px 18px 18px; display: grid; gap: 10px; }
    .gantt-row { display: grid; grid-template-columns: 260px 1fr; align-items: center; gap: 14px; min-height: 34px; }
    .gantt-name { font-size: 13px; font-weight: 900; color: #253047; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
    .gantt-track { position: relative; height: 32px; border-radius: 999px; background: linear-gradient(90deg, #eef4fb 0%, #f9fbfe 100%); overflow: visible; box-shadow: inset 0 0 0 1px rgba(226,232,240,.76); }
    .gantt-bar { position: absolute; top: 6px; height: 20px; min-width: 44px; border-radius: 999px; display: grid; place-items: center; color: #fff; font-size: 11px; font-weight: 950; box-shadow: 0 7px 16px rgba(15,23,42,.18); }
    .gantt-bar::after { content: ''; position:absolute; inset:0; border-radius: inherit; background: linear-gradient(180deg, rgba(255,255,255,.20), rgba(255,255,255,0)); pointer-events:none; }
    .gantt-star { position: absolute; top: 50%; transform: translate(8px, -50%); width: 22px; height: 22px; border-radius: 999px; display: grid; place-items: center; background: rgba(255,255,255,.96); border: 1px solid rgba(255, 138, 0, .28); font-size: 15px; line-height: 1; color: #ff8a00; box-shadow: 0 4px 12px rgba(15,23,42,.16); z-index: 3; pointer-events: none; }
    .timeline-scale .task-axis-title { text-align: left !important; color: #435168; font-weight: 950; letter-spacing: .04em; }
    .excel-preview { overflow: auto; border: 1px solid var(--line); border-radius: 14px; background: #fff; }
    .excel-preview table { min-width: 680px; width: 100%; table-layout: auto; }
    .excel-preview th, .excel-preview td { height: 36px; padding: 8px 10px; font-size: 12px; }
    .hidden { display: none !important; }
    .empty { padding: 24px; color: var(--muted); text-align: center; }

    /* v16 visual refinements */
    .brand-subtitle { display: none; }
    .brand-title { font-size: 23px; }
    .layout { grid-template-columns: minmax(0, 1fr) 232px; }
    .panel-head { padding: 18px 22px; }
    td input, td select { font-weight: 500; color: #1f2937; }
    td input { letter-spacing: 0; }
    td select {
      -webkit-appearance: none;
      appearance: none;
      background-color: #fff;
      background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='14' height='14' viewBox='0 0 20 20'%3E%3Cpath fill='%2364758B' d='M5.6 7.5a1 1 0 0 1 1.4 0l3 3 3-3a1 1 0 1 1 1.4 1.4l-3.7 3.7a1 1 0 0 1-1.4 0L5.6 8.9a1 1 0 0 1 0-1.4Z'/%3E%3C/svg%3E");
      background-repeat: no-repeat;
      background-position: right 16px center;
      background-size: 13px 13px;
      padding-right: 40px;
    }
    .date-field { font-weight: 500; letter-spacing: 0; }
    .date-field .date-placeholder { font-weight: 600; }
    .side-card { padding: 12px; }
    .side-card h3 { font-size: 13px; margin-bottom: 9px; }
    .template-list { gap: 8px; }
    .template-card { padding: 9px 10px; border-radius: 12px; }
    .template-card strong { font-size: 11px; margin-bottom: 3px; display:flex; align-items:center; gap:6px; }
    .template-card span { font-size: 10.5px; line-height: 1.35; }
    .coming-tag { display:inline-flex; align-items:center; height:18px; padding:0 6px; border-radius:999px; background:#eef5ff; color:#0f62df; font-size:9px; font-weight:900; border:1px solid #cfe0ff; white-space:nowrap; }
    .bottom-bar {
      background: linear-gradient(180deg, rgba(255,255,255,.98), rgba(243,248,255,.96));
      border-color: rgba(194, 210, 232, .95);
      box-shadow: 0 -10px 34px rgba(15, 23, 42, .12), 0 18px 45px rgba(15, 98, 223, .08);
    }
    .quick-add { background: rgba(238, 245, 255, .78); border-color:#d4e2f5; }
    .drawer-head {
      grid-template-columns: minmax(220px, 1fr) auto minmax(190px, 1fr) 40px;
      padding: 12px 18px;
    }
    .drawer-title { min-width: 0; }
    .drawer-title h2 { text-align:left; font-size:17px; }
    .drawer-tabs { justify-self: center; min-height: 42px; }
    .drawer-stats { justify-self: end; gap: 7px; }
    .compact-stat { width: 86px; height: 40px; padding: 6px 9px; border-radius: 13px; background: linear-gradient(180deg, #fff, #f8fbff); }
    .compact-stat .stat-label { font-size: 9px; color:#64748b; }
    .compact-stat .stat-value { font-size: 16px; margin-top: 4px; }
    .tab { height: 34px; min-width: 112px; }
    .gantt-card { border-radius: 18px; }
    .timeline-scale { grid-template-columns: 340px repeat(var(--week-count, 6), minmax(72px, 1fr)); padding: 12px 18px 7px; font-size: 10.5px; }
    .gantt-row { grid-template-columns: 340px 1fr; min-height: 36px; }
    .gantt-name { font-weight: 800; font-size: 12.5px; }
    .gantt-track {
      margin-right: 18px;
      background-image: repeating-linear-gradient(to right, rgba(116, 129, 150, .18) 0 1px, transparent 1px calc(100% / var(--week-count, 6))), linear-gradient(90deg, #eef4fb 0%, #f9fbfe 100%);
    }
    .gantt-bar { font-weight: 850; }

    @media (max-width: 1100px) {
      .layout { grid-template-columns: 1fr; }
      .side-stack { grid-template-columns: 1fr 1fr; }
      .project-title-box { min-width: 100%; order: 3; }
      .project-title-box input { width: 100%; min-width: 0; }
      .stage-workbench { grid-template-columns: 1fr; }
      .stage-name-control { justify-self: stretch; }
    }
    @media (max-width: 760px) {
      body { padding-bottom: 126px; }
      .app { padding: 0 14px 18px; }
      .topbar { align-items: flex-start; flex-direction: column; }
      .header-actions { width: 100%; justify-content: space-between; }
      .side-stack { grid-template-columns: 1fr; }
      .panel-head { flex-direction: column; }
      .stage-workbench { padding: 12px 14px; gap: 10px; }
      .stage-tab { min-width: 84px; height: 34px; padding: 0 12px; }
      .stage-name-control { grid-template-columns: 1fr; gap: 6px; }
      .bottom-bar { width: calc(100vw - 24px); flex-direction: column; align-items: stretch; }
      .bar-actions { display: grid; grid-template-columns: 1fr 1fr; }
      .bar-actions .btn.primary { grid-column: 1 / -1; }
      .drawer-head { grid-template-columns: 1fr 38px; align-items: start; }
      .drawer-tabs { grid-column: 1 / -1; justify-content: flex-start; order: 1; overflow-x: auto; }
      .drawer-title { grid-column: 1 / -1; order: 2; }
      .drawer-stats { grid-column: 1 / -1; order: 3; justify-content: stretch; }
      .compact-stat { flex: 1 1 0; width: auto; }
      .preview-drawer { width: 100%; height: 78vh; border-radius: 20px 20px 0 0; }
      .timeline-scale { grid-template-columns: 110px repeat(3, 1fr); }
      .timeline-scale span:nth-child(n+5) { display: none; }
      .gantt-row { grid-template-columns: 110px 1fr; }
    }
  </style>
</head>
<body>
  <div class="app">
    <header class="topbar">
      <div class="brand">
        <div class="brand-mark">G</div>
        <div>
          <h1 class="brand-title" data-i18n="appTitle">项目排期工作台</h1>
        </div>
      </div>
      <div class="header-actions">
        <div class="project-title-box">
          <label for="projectName" data-i18n="projectTitle">项目标题</label>
          <input id="projectName" type="text" placeholder="请输入项目标题">
        </div>
        <div class="segmented" aria-label="Language switcher">
          <button id="langZh" class="active" type="button">中文</button>
          <button id="langEn" type="button">English</button>
        </div>
      </div>
    </header>

    <main class="layout">
      <section class="panel task-panel">
        <div class="panel-head">
          <div>
            <h2 class="panel-title" data-i18n="taskTableTitle">任务排期表格</h2>
            <p class="panel-desc" data-i18n="taskTableDesc">像轻量 Excel 一样编辑任务；开始日期和结束日期分开填写，工作日自动计算。</p>
          </div>
        </div>
        <div id="message" class="message"></div>
        <div class="stage-workbench">
          <div class="stage-tabs" id="stageTabs" aria-label="Stage tabs"></div>
          <label class="stage-name-control">
            <span data-i18n="currentStage">当前阶段</span>
            <input id="stageNameInput" type="text" placeholder="阶段名称">
          </label>
        </div>
        <div class="table-shell">
          <div class="table-scroll">
            <table>
              <thead>
                <tr>
                  <th class="col-stage" data-i18n="stage">阶段</th>
                  <th class="col-task" data-i18n="taskName">任务</th>
                  <th class="col-stakeholder" data-i18n="stakeholder">负责人</th>
                  <th class="col-start" data-i18n="startDate">开始日期</th>
                  <th class="col-end" data-i18n="endDate">结束日期</th>
                  <th class="col-days" data-i18n="workdays">工作日</th>
                  <th class="col-status optional-status" data-i18n="status">状态</th>
                  <th class="col-actions" data-i18n="actions">操作</th>
                </tr>
              </thead>
              <tbody id="taskBody"></tbody>
            </table>
          </div>
        </div>
      </section>

      <aside class="side-stack">
        <section class="panel side-card">
          <h3 data-i18n="templatesTitle">排期模版</h3>
          <div class="template-list" id="templateList"></div>
        </section>
        <section class="panel side-card">
          <h3 data-i18n="fieldOptionsTitle">可选展示字段</h3>
          <p data-i18n="fieldOptionsDesc">阶段始终展示并导出；这里仅控制状态列。</p>
          <div class="field-controls">
            <label class="check-row"><span data-i18n="status">状态</span><input id="toggleStatus" type="checkbox"></label>
          </div>
        </section>
      </aside>
    </main>
  </div>

  <div class="bottom-bar">
    <div class="quick-add" id="quickAdd"></div>
    <div class="bar-actions">
      <input id="importFile" type="file" accept=".xlsx" hidden>
      <input id="importYear" class="import-year" type="number" min="1900" max="9999" step="1" inputmode="numeric" data-i18n-placeholder="importYearPlaceholder" aria-label="Import year">
      <button class="btn" id="importSchedule" type="button" data-i18n="importSchedule">导入排期</button>
      <button class="btn" id="addTaskButton" type="button" data-i18n="addTask">+ 新增任务</button>
      <button class="btn" id="openPreview" type="button" data-i18n="schedulePreview">排期预览</button>
      <button class="btn primary" id="exportExcel" type="button" data-i18n="exportExcel">导出 Excel</button>
    </div>
  </div>

  <div class="drawer-backdrop" id="drawerBackdrop"></div>
  <div class="calendar-popover" id="datePopover" hidden></div>

  <section class="preview-drawer" id="previewDrawer" aria-hidden="true">
    <div class="drawer-head">
      <div class="drawer-title">
        <h2 id="previewTitle" data-title-placeholder="true">项目标题</h2>
      </div>
      <div class="drawer-tabs">
        <button class="tab active" data-preview-tab="gantt" type="button" data-i18n="ganttView">甘特图</button>
        <button class="tab" data-preview-tab="excel" type="button" data-i18n="excelView">表格</button>
      </div>
      <div class="drawer-stats">
        <div class="compact-stat"><span class="stat-label" data-i18n="totalWorkweeks">总工作周</span><strong class="stat-value" id="statWeeks">0</strong></div>
        <div class="compact-stat"><span class="stat-label" data-i18n="totalWorkdays">总工作日</span><strong class="stat-value" id="statDays">0</strong></div>
      </div>
      <button class="btn drawer-close" id="closePreview" type="button">×</button>
    </div>
    <div class="drawer-body">
      <div id="ganttView" class="gantt-card">
        <div class="timeline-scale" id="timelineScale"></div>
        <div class="gantt-list" id="ganttList"></div>
      </div>
      <div id="excelView" class="excel-preview hidden"></div>
    </div>
  </section>

  <script>
    const BASE_PATH = __BASE_PATH_JSON__;
    const $ = (id) => document.getElementById(id);
    const state = { lang: 'zh', previewTab: 'gantt', tasks: [], showStatus: false, activeTemplate: 'ar', activeStage: '', dragId: null, dragStage: '' };
    const translations = {
      zh: {
        appTitle: '项目排期工作台', subtitle: '快速制作 Timeline、甘特图和 Excel 排期表，用于内部协作与客户交付。', projectTitle: '项目标题', projectPlaceholder: '请输入项目标题',
        taskTableTitle: '任务排期表格', taskTableDesc: '',
        addTask: '+ 新增任务', reset: '重置', stage: '阶段', taskName: '任务', stakeholder: '负责人', status: '状态', startDate: '开始日期', endDate: '结束日期', workdays: '工作日', actions: '操作',
        templatesTitle: '排期模版', templatesDesc: '内置模板不会因为清空缓存丢失；套用后只修改当前任务。', fieldOptionsTitle: '可选展示字段', fieldOptionsDesc: '阶段始终展示并导出；这里仅控制状态列。',
        schedulePreview: '排期预览', ganttView: '甘特图', excelView: '表格', exportExcel: '导出 Excel', importSchedule: '导入排期', importYearPlaceholder: '导入年份（可选）', importPreview: '导入预览', importSheet: '工作表', importYear: '年份', importMonthRange: '月份范围', importDateRange: '日期范围', importTaskCount: '任务数量', importValidTaskCount: '成功识别日期的任务', importConfirm: '确认导入这些任务？', dragStage: '拖动调整阶段顺序', quickAdd: '快捷添加',
        currentStage: '当前阶段', stageNamePlaceholder: '阶段名称',
        stageCount: '阶段数量', totalWorkweeks: '总工作周', totalWorkdays: '总工作日', riskItems: '风险项', unnamed: '', empty: '请添加任务并填写日期后查看排期。', invalidDate: '日期错误',
        generated: 'Timeline 已生成，你可以导出 Excel 给团队使用。', importSuccess: '排期已导入，你可以继续编辑或导出 Excel。', importError: '无法识别这个排期文件，请检查日期轴、月份标题和任务条背景色。', exportError: '请至少添加一个任务，并填写开始日期和结束日期。', rowDateError: '第 {n} 行结束日期不能早于开始日期。', nameRequired: '第 {n} 行缺少任务名称。', dateRequired: '第 {n} 行缺少开始日期或结束日期。',
        daysUnit: '{n} 天', oneDay: '1 天',
        statuses: { incomplete: '未完成', done: '已完成' },
        selectStart: '选择开始日期', selectEnd: '选择结束日期', today: '今天', clearDate: '清空', previewTaskColumn: '任务 / 事项',
        templates: { ar: ['✨ AR 项目', '需求、内容物料、内容制作、开发、UAT、上线。'], threed: ['3D 项目', '按 3D 项目标准事项自动创建完整排期。'], digital: ['数字化项目', '需求、配置、数据准备、开发联调、UAT、上线。'] }
      },
      en: {
        appTitle: 'Timeline Workbench', subtitle: 'Create Timeline, Gantt and Excel schedules for internal collaboration and client delivery.', projectTitle: 'Project Title', projectPlaceholder: 'Enter project title',
        taskTableTitle: 'Schedule Table', taskTableDesc: '',
        addTask: '+ Add task', reset: 'Reset', stage: 'Stage', taskName: 'Task', stakeholder: 'Owner', status: 'Status', startDate: 'Start Date', endDate: 'End Date', workdays: 'Workdays', actions: 'Actions',
        templatesTitle: 'Schedule Templates', templatesDesc: 'Built-in templates do not depend on browser cache. Applying one only edits current tasks.', fieldOptionsTitle: 'Optional Fields', fieldOptionsDesc: 'Stage is always shown and exported. This only controls Status.',
        schedulePreview: 'Schedule Preview', ganttView: 'Gantt', excelView: 'Table', exportExcel: 'Export Excel', importSchedule: 'Import Schedule', importYearPlaceholder: 'Import year (optional)', importPreview: 'Import preview', importSheet: 'Sheet', importYear: 'Year', importMonthRange: 'Months', importDateRange: 'Date range', importTaskCount: 'Tasks', importValidTaskCount: 'Tasks with dates', importConfirm: 'Import these tasks?', dragStage: 'Drag to reorder stages', quickAdd: 'Quick add',
        currentStage: 'Current Stage', stageNamePlaceholder: 'Stage name',
        stageCount: 'Stages', totalWorkweeks: 'Work Weeks', totalWorkdays: 'Workdays', riskItems: 'Risks', unnamed: '', empty: 'Add tasks and dates to preview the schedule.', invalidDate: 'Date error',
        generated: 'Timeline generated. You can export Excel for your team.', importSuccess: 'Schedule imported. You can keep editing or export Excel.', importError: 'Unable to read this schedule. Check the date axis, month headers, and task-bar fills.', exportError: 'Please add at least one task with start and end dates.', rowDateError: 'Row {n}: end date cannot be earlier than start date.', nameRequired: 'Row {n} is missing a task name.', dateRequired: 'Row {n} is missing a start or end date.',
        daysUnit: '{n} days', oneDay: '1 day',
        statuses: { incomplete: 'Incomplete', done: 'Done' },
        selectStart: 'Select start date', selectEnd: 'Select end date', today: 'Today', clearDate: 'Clear', previewTaskColumn: 'Task',
        templates: { ar: ['✨ AR Project', 'Requirement, assets, production, development, UAT and launch.'], threed: ['3D Project', 'Auto-create a full 3D project schedule.'], digital: ['Digitalization Project', 'Requirement, setup, data prep, development, UAT and launch.'] }
      }
    };
    const statusKeys = ['incomplete','done'];
    const ownerOptions = ['Kivisense', 'Brand', 'Brand & Kivisense'];
    const colors = { done:'#09b86f', incomplete:'#667085' };
    const ganttPalette = ['#2563eb', '#7c3aed', '#0ea5e9', '#f97316', '#16a34a', '#db2777', '#64748b', '#14b8a6'];
    const quickStages = {
      zh: ['需求', '方案', '设计', '开发', '测试', '上线', '联调'],
      en: ['Requirement', 'Proposal', 'Design', 'Development', 'Testing', 'Launch', 'Integration']
    };
    const moduleTaskTemplates = {
      zh: {
        '需求': [
          { stage: '需求', name: '需求梳理', stakeholder: 'Brand & Kivisense' },
          { stage: '需求', name: '需求确认', stakeholder: 'Brand & Kivisense' }
        ],
        '方案': [
          { stage: '方案', name: '方案设计', stakeholder: 'Kivisense' },
          { stage: '方案', name: '方案评审与确认', stakeholder: 'Brand & Kivisense' }
        ],
        '设计': [
          { stage: '设计', name: '交互设计', stakeholder: 'Kivisense' },
          { stage: '设计', name: '视觉与内容确认', stakeholder: 'Brand & Kivisense' }
        ],
        '开发': [
          { stage: '开发', name: '前后端开发', stakeholder: 'Kivisense' },
          { stage: '开发', name: '多端适配', stakeholder: 'Kivisense' },
          { stage: '开发', name: '数据埋点', stakeholder: 'Kivisense' }
        ],
        '测试': [
          { stage: '测试', name: '内部测试', stakeholder: 'Kivisense' },
          { stage: '测试', name: '测试报告', stakeholder: 'Kivisense' }
        ],
        '上线': [
          { stage: '上线', name: '上线发布', stakeholder: 'Kivisense' }
        ],
        '联调': [
          { stage: '联调', name: '系统联调', stakeholder: 'Kivisense' },
          { stage: '联调', name: '联调问题修复与确认', stakeholder: 'Brand & Kivisense' }
        ]
      },
      en: {
        'Requirement': [
          { stage: 'Requirement', name: 'Requirement review', stakeholder: 'Brand & Kivisense' },
          { stage: 'Requirement', name: 'Requirement confirmation', stakeholder: 'Brand & Kivisense' }
        ],
        'Proposal': [
          { stage: 'Proposal', name: 'Solution proposal', stakeholder: 'Kivisense' },
          { stage: 'Proposal', name: 'Proposal review and confirmation', stakeholder: 'Brand & Kivisense' }
        ],
        'Design': [
          { stage: 'Design', name: 'Interaction design', stakeholder: 'Kivisense' },
          { stage: 'Design', name: 'Visual and content confirmation', stakeholder: 'Brand & Kivisense' }
        ],
        'Development': [
          { stage: 'Development', name: 'Frontend and backend development', stakeholder: 'Kivisense' },
          { stage: 'Development', name: 'Multi-device adaptation', stakeholder: 'Kivisense' },
          { stage: 'Development', name: 'Data tracking', stakeholder: 'Kivisense' }
        ],
        'Testing': [
          { stage: 'Testing', name: 'Internal testing', stakeholder: 'Kivisense' },
          { stage: 'Testing', name: 'Test report', stakeholder: 'Kivisense' }
        ],
        'Launch': [
          { stage: 'Launch', name: 'Launch release', stakeholder: 'Kivisense' }
        ],
        'Integration': [
          { stage: 'Integration', name: 'System integration', stakeholder: 'Kivisense' },
          { stage: 'Integration', name: 'Integration fixes and confirmation', stakeholder: 'Brand & Kivisense' }
        ]
      }
    };
    function buildModuleTasks(lang, keys) {
      const source = moduleTaskTemplates[lang] || moduleTaskTemplates.zh;
      return keys.flatMap(key => (source[key] || []).map(item => ({ ...item, status: 'incomplete', start: '', end: '' })));
    }
    const templateTasks = {
      zh: {
        ar: [
          { stage: '需求', name: '需求梳理与范围确认', stakeholder: 'Kivisense', status: 'incomplete', start: '2026-06-10', end: '2026-06-10' },
          { stage: '方案', name: 'AR 方案确认', stakeholder: 'Brand & Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: '设计', name: '交互与视觉设计', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: '开发', name: '程序开发与多端适配', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: '联调', name: '系统联调与数据埋点', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: '测试', name: 'UAT 验收测试', stakeholder: 'Brand & Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: '上线', name: '上线发布', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' }
        ],
        threed: buildModuleTasks('zh', quickStages.zh),
        digital: [
          { stage: '需求', name: '业务需求确认', stakeholder: 'Brand & Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: '配置', name: '后台配置与规则确认', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: '数据', name: '数据准备与校验', stakeholder: 'Brand', status: 'incomplete', start: '', end: '' },
          { stage: '开发', name: '开发联调', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: '联调', name: '系统联调', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: '测试', name: 'UAT 验收', stakeholder: 'Brand & Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: '上线', name: '上线发布', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' }
        ]
      },
      en: {
        ar: [
          { stage: 'Requirement', name: 'Requirement review and scope confirmation', stakeholder: 'Kivisense', status: 'incomplete', start: '2026-06-10', end: '2026-06-10' },
          { stage: 'Proposal', name: 'AR proposal confirmation', stakeholder: 'Brand & Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: 'Design', name: 'Interaction and visual design', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: 'Development', name: 'Development and device adaptation', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: 'Integration', name: 'System integration and data tracking', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: 'Testing', name: 'UAT acceptance testing', stakeholder: 'Brand & Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: 'Launch', name: 'Launch release', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' }
        ],
        threed: buildModuleTasks('en', quickStages.en),
        digital: [
          { stage: 'Requirement', name: 'Business requirement confirmation', stakeholder: 'Brand & Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: 'Setup', name: 'Admin setup and rule confirmation', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: 'Data', name: 'Data preparation and validation', stakeholder: 'Brand', status: 'incomplete', start: '', end: '' },
          { stage: 'Development', name: 'Development integration', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: 'Integration', name: 'System integration', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: 'Testing', name: 'UAT acceptance', stakeholder: 'Brand & Kivisense', status: 'incomplete', start: '', end: '' },
          { stage: 'Launch', name: 'Launch release', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' }
        ]
      }
    };
    const templates = [{ id: 'ar' }, { id: 'threed' }, { id: 'digital' }];
    function getTemplateTasks(id) {
      const tasksByLang = templateTasks[state.lang] || templateTasks.zh;
      return tasksByLang[id] || tasksByLang.ar;
    }
    function defaultStageName() { return state.lang === 'zh' ? '未分类' : 'Uncategorized'; }
    function normalizeStage(value) { return String(value ?? '').trim() || defaultStageName(); }
    function appPath(path) { if (!BASE_PATH) return path; return BASE_PATH + path; }
    function t(key) { return translations[state.lang][key] || key; }
    function uid() { return 'task-' + Math.random().toString(16).slice(2) + Date.now().toString(16); }
    function escapeHtml(value) { return String(value ?? '').replace(/[&<>'"]/g, ch => ({'&':'&amp;','<':'&lt;','>':'&gt;',"'":'&#39;','"':'&quot;'}[ch])); }
    function withReplacements(text, values) { return text.replace(/\{(\w+)\}/g, (_, key) => values[key] ?? ''); }
    function parseDate(value) { if (!value) return null; const d = new Date(value + 'T00:00:00'); return Number.isNaN(d.getTime()) ? null : d; }
    function isInvalidRange(task) { const s = parseDate(task.start); const e = parseDate(task.end); return !!(s && e && e < s); }
    function countWorkdays(startValue, endValue) {
      const start = parseDate(startValue), end = parseDate(endValue);
      if (!start || !end || end < start) return 0;
      let count = 0; const cursor = new Date(start);
      while (cursor <= end) { const day = cursor.getDay(); if (day !== 0 && day !== 6) count++; cursor.setDate(cursor.getDate() + 1); }
      return count;
    }
    function daysLabel(n) { return n === 1 ? t('oneDay') : withReplacements(t('daysUnit'), { n }); }
    function formatDate(value) {
      if (!value) return '';
      const d = parseDate(value); if (!d) return value;
      if (state.lang === 'zh') return value.replaceAll('-', '/');
      return d.toLocaleDateString('en-US', { month: 'short', day: '2-digit', year: 'numeric' });
    }
    function formatRange(task) { return task.start && task.end ? `${formatDate(task.start)} - ${formatDate(task.end)}` : ''; }
    function showMessage(text, type='success') {
      const el = $('message'); el.textContent = text; el.className = `message show ${type}`;
      clearTimeout(showMessage.timer); showMessage.timer = setTimeout(() => el.className = 'message', 3200);
    }

    function toISODate(day) {
      const y = day.getFullYear();
      const m = String(day.getMonth() + 1).padStart(2, '0');
      const d = String(day.getDate()).padStart(2, '0');
      return `${y}-${m}-${d}`;
    }
    function openDatePicker(taskId, field, anchor) {
      const task = state.tasks.find(x => x.id === taskId);
      if (!task) return;
      const selected = parseDate(task[field]);
      const month = selected || new Date();
      state.datePicker = { taskId, field, selected: task[field] || '', month: new Date(month.getFullYear(), month.getMonth(), 1) };
      renderDatePicker(anchor);
    }
    function closeDatePicker() {
      $('datePopover').hidden = true;
      state.datePicker = null;
    }
    function shiftPickerMonth(delta) {
      if (!state.datePicker) return;
      const month = state.datePicker.month;
      state.datePicker.month = new Date(month.getFullYear(), month.getMonth() + delta, 1);
      renderDatePicker(state.datePicker.anchor);
    }
    function renderDatePicker(anchor) {
      const picker = state.datePicker;
      if (!picker) return;
      picker.anchor = anchor || picker.anchor;
      const pop = $('datePopover');
      const month = picker.month;
      const title = state.lang === 'zh' ? `${month.getFullYear()} 年 ${month.getMonth() + 1} 月` : month.toLocaleDateString('en-US', { month: 'long', year: 'numeric' });
      const week = state.lang === 'zh' ? ['日','一','二','三','四','五','六'] : ['S','M','T','W','T','F','S'];
      const first = new Date(month.getFullYear(), month.getMonth(), 1);
      const start = new Date(first); start.setDate(first.getDate() - first.getDay());
      const today = toISODate(new Date());
      const days = [];
      for (let i = 0; i < 42; i++) {
        const d = new Date(start); d.setDate(start.getDate() + i);
        const iso = toISODate(d);
        const classes = ['calendar-day'];
        if (d.getMonth() !== month.getMonth()) classes.push('muted');
        if (iso === today) classes.push('today');
        if (iso === picker.selected) classes.push('selected');
        days.push(`<button type="button" class="${classes.join(' ')}" data-date-value="${iso}">${d.getDate()}</button>`);
      }
      pop.innerHTML = `<div class="calendar-head"><div class="calendar-title">${title}</div><div class="calendar-nav"><button type="button" data-cal-prev>‹</button><button type="button" data-cal-next>›</button></div></div><div class="calendar-week">${week.map(x => `<span>${x}</span>`).join('')}</div><div class="calendar-grid">${days.join('')}</div><div class="calendar-foot"><button type="button" data-cal-clear>${t('clearDate')}</button><button type="button" class="primary-lite" data-cal-today>${t('today')}</button></div>`;
      pop.hidden = false;
      pop.onmousedown = (event) => event.stopPropagation();
      pop.onclick = (event) => event.stopPropagation();
      const rect = picker.anchor.getBoundingClientRect();
      const top = Math.min(window.innerHeight - pop.offsetHeight - 12, rect.bottom + 8);
      const left = Math.min(window.innerWidth - pop.offsetWidth - 12, Math.max(12, rect.left));
      pop.style.top = `${Math.max(12, top)}px`;
      pop.style.left = `${left}px`;
      pop.querySelector('[data-cal-prev]').addEventListener('click', (event) => { event.preventDefault(); event.stopPropagation(); shiftPickerMonth(-1); });
      pop.querySelector('[data-cal-next]').addEventListener('click', (event) => { event.preventDefault(); event.stopPropagation(); shiftPickerMonth(1); });
      pop.querySelectorAll('[data-cal-today]').forEach(btn => btn.addEventListener('click', (event) => {
        event.stopPropagation();
        const now = new Date(); picker.month = new Date(now.getFullYear(), now.getMonth(), 1); picker.selected = toISODate(now); updateTask(picker.taskId, { [picker.field]: picker.selected }); closeDatePicker();
      }));
      pop.querySelector('[data-cal-clear]').addEventListener('click', (event) => { event.stopPropagation(); updateTask(picker.taskId, { [picker.field]: '' }); closeDatePicker(); });
      pop.querySelectorAll('[data-date-value]').forEach(btn => btn.addEventListener('click', (event) => { event.stopPropagation(); updateTask(picker.taskId, { [picker.field]: btn.dataset.dateValue }); closeDatePicker(); }));
    }

    function emptyTask(stage) {
      return { id: uid(), stage: normalizeStage(stage), name: '', stakeholder: 'Kivisense', status: 'incomplete', start: '', end: '' };
    }
    function cloneTask(task) {
      return {
        id: uid(),
        stage: normalizeStage(task.stage || task.model),
        name: task.name || '',
        stakeholder: task.stakeholder || 'Kivisense',
        status: task.status || 'incomplete',
        start: task.start || '',
        end: task.end || ''
      };
    }
    function ensureTaskStages() {
      if (!state.tasks.length) state.tasks.push(emptyTask(state.activeStage || defaultStageName()));
      state.tasks.forEach(task => { task.stage = normalizeStage(task.stage); });
    }
    function getStageOrder() {
      ensureTaskStages();
      const order = [];
      state.tasks.forEach(task => {
        const stage = normalizeStage(task.stage);
        task.stage = stage;
        if (!order.includes(stage)) order.push(stage);
      });
      return order;
    }
    function getActiveStage() {
      const stages = getStageOrder();
      if (!state.activeStage || !stages.includes(state.activeStage)) state.activeStage = stages[0] || defaultStageName();
      return state.activeStage;
    }
    function activeStageTasks() {
      const active = getActiveStage();
      return state.tasks.filter(task => normalizeStage(task.stage) === active);
    }
    function applyTemplate(id) {
      const template = templates.find(x => x.id === id) || templates[0];
      state.activeTemplate = template.id;
      state.tasks = getTemplateTasks(template.id).map(cloneTask);
      state.activeStage = state.tasks[0]?.stage || defaultStageName();
      renderAll();
    }
    function addTask(stage='') {
      const targetStage = normalizeStage(stage || state.activeStage || getActiveStage());
      state.tasks.push(emptyTask(targetStage));
      state.activeStage = targetStage;
      renderAll();
    }
    function addModuleTasks(stage) {
      const targetStage = normalizeStage(stage);
      const templatesForLang = moduleTaskTemplates[state.lang] || {};
      const items = templatesForLang[targetStage] || [{ stage: targetStage, name: targetStage, stakeholder: 'Kivisense' }];
      state.tasks.push(...items.map(item => ({ id: uid(), stage: normalizeStage(item.stage || targetStage), name: item.name || targetStage, stakeholder: item.stakeholder || 'Kivisense', status: 'incomplete', start: '', end: '' })));
      state.activeStage = targetStage;
      renderAll();
    }
    function renameActiveStage(value) {
      const currentStage = getActiveStage();
      const nextStage = normalizeStage(value);
      state.tasks.forEach(task => {
        if (normalizeStage(task.stage) === currentStage) task.stage = nextStage;
      });
      state.activeStage = nextStage;
      renderAll();
    }
    function updateTask(id, patch) {
      const task = state.tasks.find(x => x.id === id);
      if (task) {
        if (Object.prototype.hasOwnProperty.call(patch, 'stage')) patch.stage = normalizeStage(patch.stage);
        Object.assign(task, patch);
      }
      renderAll(false);
    }
    function clearDragIndicators() { document.querySelectorAll('#taskBody tr').forEach(row => row.classList.remove('drag-over-before', 'drag-over-after')); }
    function moveTask(id, direction) {
      const visible = activeStageTasks();
      const index = visible.findIndex(x => x.id === id);
      const target = visible[index + direction];
      if (!target) return;
      reorderTask(id, target.id, direction < 0 ? 'before' : 'after');
    }
    function reorderTask(dragId, targetId, position='before') {
      const from = state.tasks.findIndex(x => x.id === dragId);
      let to = state.tasks.findIndex(x => x.id === targetId);
      if (from < 0 || to < 0 || from === to) return;
      const [item] = state.tasks.splice(from, 1);
      if (from < to) to -= 1;
      if (position === 'after') to += 1;
      state.tasks.splice(Math.max(0, Math.min(to, state.tasks.length)), 0, item);
      renderAll();
    }
    function reorderStageTaskBlocks(tasks, stageOrder) {
      const buckets = new Map(stageOrder.map(stage => [stage, []]));
      tasks.forEach(task => {
        const stage = normalizeStage(task.stage);
        task.stage = stage;
        if (!buckets.has(stage)) buckets.set(stage, []);
        buckets.get(stage).push(task);
      });
      return stageOrder.flatMap(stage => buckets.get(stage) || []);
    }
    function clearStageDragIndicators() {
      document.querySelectorAll('#stageTabs .stage-tab').forEach(tab => tab.classList.remove('stage-dragging', 'stage-drag-over-before', 'stage-drag-over-after'));
    }
    function reorderStages(dragStage, targetStage, position='before') {
      const stages = getStageOrder();
      const from = stages.indexOf(dragStage);
      let to = stages.indexOf(targetStage);
      if (from < 0 || to < 0 || from === to) return;
      const [stage] = stages.splice(from, 1);
      if (from < to) to -= 1;
      if (position === 'after') to += 1;
      stages.splice(Math.max(0, Math.min(to, stages.length)), 0, stage);
      state.tasks = reorderStageTaskBlocks(state.tasks, stages);
      renderAll();
    }
    function removeTask(id) { state.tasks = state.tasks.filter(x => x.id !== id); renderAll(); }
    function duplicateTask(id) { const i = state.tasks.findIndex(x => x.id === id); if (i < 0) return; const copy = cloneTask(state.tasks[i]); state.tasks.splice(i + 1, 0, copy); state.activeStage = copy.stage; renderAll(); }
    function statusLabel(status) { return translations[state.lang].statuses[status] || status; }
    function renderTemplates() {
      $('templateList').innerHTML = templates.map(template => {
        const info = translations[state.lang].templates[template.id];
        const soon = template.id === 'ar' || template.id === 'digital';
        const badge = soon ? `<em class="coming-tag">${state.lang === 'zh' ? '敬请期待' : 'Soon'}</em>` : '';
        return `<button type="button" class="template-card ${state.activeTemplate === template.id ? 'active' : ''}" data-template="${template.id}"><strong>${info[0]}${badge}</strong><span>${info[1]}</span></button>`;
      }).join('');
      document.querySelectorAll('[data-template]').forEach(btn => btn.addEventListener('click', () => applyTemplate(btn.dataset.template)));
    }
    function renderQuickAdd() {
      const stages = quickStages[state.lang] || quickStages.zh;
      $('quickAdd').innerHTML = `<span class="quick-add-label">${t('quickAdd')}</span>` + stages.map(stage => `<button class="btn small" type="button" data-quick="${escapeHtml(stage)}">+ ${escapeHtml(stage)}</button>`).join('');
      document.querySelectorAll('[data-quick]').forEach(btn => btn.addEventListener('click', () => addModuleTasks(btn.dataset.quick)));
    }
    function renderStageTabs() {
      const stages = getStageOrder();
      const active = getActiveStage();
      const counts = stages.reduce((acc, stage) => {
        acc[stage] = state.tasks.filter(task => normalizeStage(task.stage) === stage).length;
        return acc;
      }, {});
      $('stageTabs').innerHTML = stages.map((stage, index) => `<button type="button" draggable="true" class="stage-tab ${stage === active ? 'active' : ''}" data-stage-index="${index}" data-stage="${escapeHtml(stage)}" title="${escapeHtml(t('dragStage'))}"><span>${escapeHtml(stage)}</span><span class="stage-tab-count">${counts[stage] || 0}</span></button>`).join('');
      document.querySelectorAll('[data-stage-index]').forEach(btn => {
        btn.addEventListener('click', () => {
        state.activeStage = stages[Number(btn.dataset.stageIndex)];
        renderAll();
        });
        btn.addEventListener('dragstart', event => {
          state.dragStage = btn.dataset.stage;
          btn.classList.add('stage-dragging');
          event.dataTransfer.effectAllowed = 'move';
          event.dataTransfer.setData('text/plain', state.dragStage);
        });
        btn.addEventListener('dragend', () => {
          state.dragStage = '';
          clearStageDragIndicators();
        });
        btn.addEventListener('dragover', event => {
          const dragStage = event.dataTransfer.getData('text/plain') || state.dragStage;
          if (!dragStage || dragStage === btn.dataset.stage) return;
          event.preventDefault();
          event.dataTransfer.dropEffect = 'move';
          clearStageDragIndicators();
          const rect = btn.getBoundingClientRect();
          btn.classList.add(event.clientX < rect.left + rect.width / 2 ? 'stage-drag-over-before' : 'stage-drag-over-after');
        });
        btn.addEventListener('dragleave', () => btn.classList.remove('stage-drag-over-before', 'stage-drag-over-after'));
        btn.addEventListener('drop', event => {
          event.preventDefault();
          const dragStage = event.dataTransfer.getData('text/plain') || state.dragStage;
          const position = btn.classList.contains('stage-drag-over-after') ? 'after' : 'before';
          clearStageDragIndicators();
          reorderStages(dragStage, btn.dataset.stage, position);
        });
      });
      const input = $('stageNameInput');
      input.value = active;
      input.placeholder = t('stageNamePlaceholder');
      input.onchange = () => renameActiveStage(input.value);
      input.onkeydown = (event) => { if (event.key === 'Enter') input.blur(); };
    }
    function renderTasks(rebind=true) {
      const tasks = activeStageTasks();
      $('taskBody').innerHTML = tasks.map((task) => {
        const days = countWorkdays(task.start, task.end);
        const invalid = isInvalidRange(task);
        return `<tr data-id="${task.id}">
          <td class="col-stage"><input data-field="stage" value="${escapeHtml(task.stage)}" placeholder="${t('stage')}"></td>
          <td class="col-task"><input data-field="name" value="${escapeHtml(task.name)}" placeholder="${t('taskName')}"></td>
          <td class="col-stakeholder"><select data-field="stakeholder">${ownerOptions.map(opt => `<option value="${opt}" ${opt === task.stakeholder ? 'selected' : ''}>${escapeHtml(opt)}</option>`).join('')}</select></td>
          <td class="col-start"><button class="date-field" type="button" data-date-field="start"><span class="${task.start ? '' : 'date-placeholder'}">${task.start ? formatDate(task.start) : t('selectStart')}</span><span class="date-icon">▾</span></button></td>
          <td class="col-end"><button class="date-field" type="button" data-date-field="end"><span class="${task.end ? '' : 'date-placeholder'}">${task.end ? formatDate(task.end) : t('selectEnd')}</span><span class="date-icon">▾</span></button></td>
          <td class="col-days"><span class="days-pill ${invalid ? 'error' : ''}">${invalid ? t('invalidDate') : daysLabel(days)}</span></td>
          <td class="col-status optional-status"><span class="status-wrap"><span class="status-chip ${task.status}">${statusLabel(task.status)}</span><select data-field="status">${statusKeys.map(key => `<option value="${key}" ${key === task.status ? 'selected' : ''}>${statusLabel(key)}</option>`).join('')}</select></span></td>
          <td class="col-actions"><div class="row-actions"><button class="icon-btn drag-handle" draggable="true" data-action="drag" title="Drag">↕</button><button class="icon-btn" data-action="duplicate" title="Duplicate">⧉</button><button class="icon-btn" data-action="up" title="Up">↑</button><button class="icon-btn" data-action="down" title="Down">↓</button><button class="icon-btn" data-action="delete" title="Delete">×</button></div></td>
        </tr>`;
      }).join('');
      applyFieldVisibility();
      bindRowEvents();
    }
    function bindRowEvents() {
      document.querySelectorAll('#taskBody tr').forEach(row => {
        const id = row.dataset.id;
        row.querySelectorAll('input, select').forEach(input => input.addEventListener('change', () => {
          updateTask(id, { [input.dataset.field]: input.value });
        }));
        row.querySelectorAll('input[type="text"], input:not([type])').forEach(input => input.addEventListener('input', () => {
          const task = state.tasks.find(x => x.id === id);
          if (task) task[input.dataset.field] = input.value;
          renderPreview();
        }));
        row.querySelectorAll('[data-date-field]').forEach(btn => btn.addEventListener('click', () => openDatePicker(id, btn.dataset.dateField, btn)));
        const dragHandle = row.querySelector('.drag-handle');
        if (dragHandle) {
          dragHandle.addEventListener('dragstart', event => { state.dragId = id; row.classList.add('dragging'); event.dataTransfer.effectAllowed = 'move'; event.dataTransfer.setData('text/plain', id); });
          dragHandle.addEventListener('dragend', () => { state.dragId = null; row.classList.remove('dragging'); clearDragIndicators(); });
        }
        row.addEventListener('dragover', event => {
          if (state.dragId && state.dragId !== id) {
            event.preventDefault(); event.dataTransfer.dropEffect = 'move';
            clearDragIndicators();
            const rect = row.getBoundingClientRect();
            row.classList.add(event.clientY < rect.top + rect.height / 2 ? 'drag-over-before' : 'drag-over-after');
          }
        });
        row.addEventListener('dragleave', () => row.classList.remove('drag-over-before', 'drag-over-after'));
        row.addEventListener('drop', event => {
          event.preventDefault();
          const dragId = event.dataTransfer.getData('text/plain') || state.dragId;
          const position = row.classList.contains('drag-over-after') ? 'after' : 'before';
          clearDragIndicators();
          reorderTask(dragId, id, position);
        });
        row.querySelectorAll('[data-action]').forEach(btn => btn.addEventListener('click', () => {
          const action = btn.dataset.action;
          if (action === 'drag') return;
          if (action === 'duplicate') duplicateTask(id);
          if (action === 'up') moveTask(id, -1);
          if (action === 'down') moveTask(id, 1);
          if (action === 'delete') removeTask(id);
        }));
      });
    }
    function applyFieldVisibility() {
      document.querySelectorAll('.optional-status').forEach(el => el.classList.toggle('hidden', !state.showStatus));
    }
    function renderPreview() {
      const valid = state.tasks.filter(task => task.start && task.end && !isInvalidRange(task));
      const totalDays = valid.reduce((sum, task) => sum + countWorkdays(task.start, task.end), 0);
      const totalWeeks = totalDays ? Math.ceil(totalDays / 5) : 0;
      const projectTitle = $('projectName').value.trim();
      $('previewTitle').textContent = projectTitle || t('projectTitle');
      $('statWeeks').textContent = totalWeeks;
      $('statDays').textContent = totalDays;
      renderGantt(valid);
      renderExcel();
    }
    function startOfWeek(day) {
      const d = new Date(day);
      const delta = (d.getDay() + 6) % 7;
      d.setDate(d.getDate() - delta);
      return d;
    }
    function endOfWeek(day) {
      const d = startOfWeek(day);
      d.setDate(d.getDate() + 6);
      return d;
    }
    function weekLabel(day) {
      return state.lang === 'zh'
        ? `${day.getMonth()+1}/${String(day.getDate()).padStart(2,'0')}`
        : day.toLocaleDateString('en-US', { month: 'short', day: '2-digit' });
    }
    function renderGantt(tasks) {
      if (!tasks.length) { $('timelineScale').innerHTML = ''; $('ganttList').innerHTML = `<div class="empty">${t('empty')}</div>`; return; }
      const starts = tasks.map(x => parseDate(x.start)); const ends = tasks.map(x => parseDate(x.end));
      const firstDay = new Date(Math.min(...starts)); const lastDay = new Date(Math.max(...ends));
      const min = startOfWeek(firstDay); const max = endOfWeek(lastDay);
      const span = Math.max(1, Math.round((max - min) / 86400000) + 1);
      const weekCount = Math.max(1, Math.ceil(span / 7));
      const weekLabels = [];
      for (let i = 0; i < weekCount; i++) { const d = new Date(min); d.setDate(min.getDate() + i * 7); weekLabels.push(weekLabel(d)); }
      $('timelineScale').style.setProperty('--week-count', weekCount);
      $('timelineScale').innerHTML = `<span class="task-axis-title">${t('previewTaskColumn')}</span>` + weekLabels.map(x => `<span>${x}</span>`).join('');
      $('ganttList').innerHTML = tasks.map((task, index) => {
        const s = parseDate(task.start), e = parseDate(task.end);
        const left = Math.max(0, ((s - min) / 86400000) / span * 100);
        const width = Math.max(3.2, (((e - s) / 86400000) + 1) / span * 100);
        const endPct = Math.min(100, left + width);
        const days = countWorkdays(task.start, task.end);
        const name = task.name;
        const barColor = task.status === 'done' ? colors.done : ganttPalette[index % ganttPalette.length];
        return `<div class="gantt-row"><div class="gantt-name" title="${escapeHtml(task.name)}">${escapeHtml(name)}</div><div class="gantt-track" style="--week-count:${weekCount}"><div class="gantt-bar" style="left:${left}%;width:${width}%;background:${barColor}">${days}d</div><span class="gantt-star" style="left:${endPct}%">★</span></div></div>`;
      }).join('');
    }
    function renderExcel() {
      const headers = [t('stage')];
      headers.push(t('taskName'), t('stakeholder'));
      if (state.showStatus) headers.push(t('status'));
      headers.push(t('startDate'), t('endDate'), t('workdays'));
      const rows = state.tasks.map(task => {
        const cells = [normalizeStage(task.stage)];
        cells.push(task.name, task.stakeholder);
        if (state.showStatus) cells.push(statusLabel(task.status));
        cells.push(formatDate(task.start), formatDate(task.end), daysLabel(countWorkdays(task.start, task.end)));
        return cells;
      });
      $('excelView').innerHTML = `<table><thead><tr>${headers.map(h => `<th>${escapeHtml(h)}</th>`).join('')}</tr></thead><tbody>${rows.map(row => `<tr>${row.map(cell => `<td>${escapeHtml(cell)}</td>`).join('')}</tr>`).join('')}</tbody></table>`;
    }
    function validate() {
      if (!state.tasks.length) throw new Error(t('exportError'));
      state.tasks.forEach((task, index) => {
        task.stage = normalizeStage(task.stage);
        if (!task.name.trim()) throw new Error(withReplacements(t('nameRequired'), { n: index + 1 }));
        if (!task.start || !task.end) throw new Error(withReplacements(t('dateRequired'), { n: index + 1 }));
        if (isInvalidRange(task)) throw new Error(withReplacements(t('rowDateError'), { n: index + 1 }));
      });
    }
    function stakeholderToOwners(value) {
      if (value === 'Brand') return ['Brands'];
      if (value === 'Brand & Kivisense') return ['Kivisense', 'Brands'];
      return ['Kivisense'];
    }
    function ownersToStakeholder(owners) {
      const list = Array.isArray(owners) ? owners.map(x => String(x).toLowerCase()) : [];
      const hasKivisense = list.some(x => x.includes('kivisense') || x.includes('弥知') || x.includes('我方'));
      const hasBrand = list.some(x => x.includes('brand') || x.includes('品牌') || x.includes('客户'));
      if (hasKivisense && hasBrand) return 'Brand & Kivisense';
      if (hasBrand) return 'Brand';
      return 'Kivisense';
    }
    function applyImportedSchedule(data) {
      const importedTasks = Array.isArray(data.tasks) ? data.tasks : [];
      if (!importedTasks.length) throw new Error(t('importError'));
      $('projectName').value = data.project_name || '';
      state.tasks = importedTasks.map(task => ({
        id: uid(),
        stage: normalizeStage(task.model || task.stage),
        name: task.name || '',
        stakeholder: ownersToStakeholder(task.owners),
        status: task.status === 'done' ? 'done' : 'incomplete',
        start: task.start || '',
        end: task.end || ''
      }));
      state.activeStage = state.tasks[0]?.stage || defaultStageName();
      state.showStatus = !!data.include_status;
      $('toggleStatus').checked = state.showStatus;
      renderAll();
      showMessage(t('importSuccess'));
    }
    function formatImportPreview(data) {
      const preview = data.import_preview || {};
      const inferredYear = preview.year_inferred
        ? (state.lang === 'zh' ? `（推断，来源：${preview.inferred_year_source}，需确认）` : ` (inferred from ${preview.inferred_year_source}; confirm)`)
        : '';
      const lines = [
        t('importPreview'),
        `${t('importSheet')}: ${preview.sheet || ''}`,
        `${t('importYear')}: ${preview.year || ''}${inferredYear}`,
        `${t('importMonthRange')}: ${(preview.months || []).join(', ')}`,
        `${t('importDateRange')}: ${preview.date_range || ''}`,
        `${t('importTaskCount')}: ${preview.task_count || 0}`,
        `${t('importValidTaskCount')}: ${preview.valid_date_task_count || 0}`
      ];
      (preview.warnings || []).forEach(warning => lines.push(`- ${warning}`));
      lines.push('', t('importConfirm'));
      return lines.join('\n');
    }
    function fileToBase64(file) {
      return new Promise((resolve, reject) => {
        const reader = new FileReader();
        reader.onload = () => resolve(String(reader.result).split(',')[1] || '');
        reader.onerror = () => reject(reader.error || new Error('File read failed'));
        reader.readAsDataURL(file);
      });
    }
    async function importScheduleFile(file) {
      if (!file) return;
      try {
        const file_b64 = await fileToBase64(file);
        const importYear = $('importYear').value.trim();
        const response = await fetch(appPath('/import'), { method: 'POST', headers: { 'Content-Type': 'application/json' }, body: JSON.stringify({ filename: file.name, file_b64, import_year: importYear || null }) });
        const data = await response.json().catch(() => ({}));
        if (!response.ok) throw new Error(data.error || t('importError'));
        if (!window.confirm(formatImportPreview(data))) return;
        applyImportedSchedule(data);
      } catch (err) {
        showMessage(err.message || t('importError'), 'error');
      } finally {
        $('importFile').value = '';
      }
    }
    function collectExportTasks() {
      validate();
      return state.tasks.map(task => ({ model: normalizeStage(task.stage), name: task.name, owners: stakeholderToOwners(task.stakeholder), status: task.status === 'done' ? 'done' : 'incomplete', start: task.start, end: task.end, workdays: countWorkdays(task.start, task.end) }));
    }
    async function exportExcel() {
      try {
        const tasks = collectExportTasks();
        const response = await fetch(appPath('/generate'), { method: 'POST', headers: { 'Content-Type': 'application/json' }, body: JSON.stringify({ project_name: $('projectName').value.trim() || 'Timeline', tasks, include_model: true, include_status: state.showStatus, language: state.lang }) });
        if (!response.ok) { const problem = await response.json().catch(() => ({ error: 'Export failed' })); throw new Error(problem.error || 'Export failed'); }
        const blob = await response.blob(); const url = URL.createObjectURL(blob);
        const safeName = ($('projectName').value || 'timeline').replace(/[\\/:*?"<>|\s]+/g, '_');
        const a = document.createElement('a'); a.href = url; a.download = `${safeName}_timeline.xlsx`; document.body.appendChild(a); a.click(); a.remove(); URL.revokeObjectURL(url);
        showMessage(t('generated'));
      } catch (err) { showMessage(err.message, 'error'); }
    }
    function setLanguage(lang) {
      state.lang = lang;
      document.documentElement.lang = lang === 'zh' ? 'zh-CN' : 'en';
      $('langZh').classList.toggle('active', lang === 'zh'); $('langEn').classList.toggle('active', lang === 'en');
      document.querySelectorAll('[data-i18n]').forEach(el => { el.textContent = t(el.dataset.i18n); });
      document.querySelectorAll('[data-i18n-placeholder]').forEach(el => { el.placeholder = t(el.dataset.i18nPlaceholder); });
      $('projectName').placeholder = t('projectPlaceholder');
      $('stageNameInput').placeholder = t('stageNamePlaceholder');
      renderAll();
    }
    function renderAll() { ensureTaskStages(); renderTemplates(); renderQuickAdd(); renderStageTabs(); renderTasks(); renderPreview(); }
    function openDrawer() { renderPreview(); $('drawerBackdrop').classList.add('open'); $('previewDrawer').classList.add('open'); $('previewDrawer').setAttribute('aria-hidden', 'false'); }
    function closeDrawer() { $('drawerBackdrop').classList.remove('open'); $('previewDrawer').classList.remove('open'); $('previewDrawer').setAttribute('aria-hidden', 'true'); }
    $('langZh').addEventListener('click', () => setLanguage('zh'));
    $('langEn').addEventListener('click', () => setLanguage('en'));
    $('projectName').addEventListener('input', renderPreview);
    $('importSchedule').addEventListener('click', () => $('importFile').click());
    $('importFile').addEventListener('change', e => importScheduleFile(e.target.files && e.target.files[0]));
    $('addTaskButton').addEventListener('click', () => addTask(''));
    $('exportExcel').addEventListener('click', exportExcel);
    $('openPreview').addEventListener('click', openDrawer);
    $('closePreview').addEventListener('click', closeDrawer);
    $('drawerBackdrop').addEventListener('click', closeDrawer);
    document.addEventListener('click', event => { if (!event.target.closest('.calendar-popover') && !event.target.closest('[data-date-field]')) closeDatePicker(); });
    window.addEventListener('resize', closeDatePicker);
    window.addEventListener('scroll', closeDatePicker, true);
    $('toggleStatus').addEventListener('change', e => { state.showStatus = e.target.checked; renderAll(); });
    document.querySelectorAll('[data-preview-tab]').forEach(tab => tab.addEventListener('click', () => { state.previewTab = tab.dataset.previewTab; document.querySelectorAll('[data-preview-tab]').forEach(item => item.classList.toggle('active', item === tab)); $('ganttView').classList.toggle('hidden', state.previewTab !== 'gantt'); $('excelView').classList.toggle('hidden', state.previewTab !== 'excel'); renderPreview(); }));
    $('toggleStatus').checked = state.showStatus; applyTemplate('ar'); setLanguage('zh');
  </script>
</body>
</html>
"""


DATE_RE = re.compile(r"\b\d{4}[-/]\d{1,2}[-/]\d{1,2}\b")
DURATION_RE = re.compile(r"\b\d+\s*(?:天|day|days|workday|workdays|个工作日)\b", re.IGNORECASE)
OWNER_RE = re.compile(r"kivisense|kv|brand|brands|client|我方|客户|品牌方", re.IGNORECASE)



def normalize_text(value) -> str:
    return str(value or "").strip()


def cell_text(value) -> str:
    return normalize_text(value).replace("\n", " ")


def month_from_label(value) -> int | None:
    text = cell_text(value).lower()
    if not text:
        return None
    zh_match = re.search(r"(\d{1,2})\s*月", text)
    if zh_match:
        month = int(zh_match.group(1))
        return month if 1 <= month <= 12 else None
    zh_months = {
        "一月": 1, "二月": 2, "三月": 3, "四月": 4, "五月": 5, "六月": 6,
        "七月": 7, "八月": 8, "九月": 9, "十月": 10, "十一月": 11, "十二月": 12,
    }
    if text in zh_months:
        return zh_months[text]
    numeric = re.fullmatch(r"\d{1,2}", text)
    if numeric:
        month = int(text)
        return month if 1 <= month <= 12 else None
    months = {
        "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
        "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6,
        "jul": 7, "july": 7, "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
    }
    return months.get(text)


class ScheduleImportError(ValueError):
    """A user-facing import error with stable details for the HTTP API."""

    def __init__(self, code: str, message: str, details: dict | None = None):
        super().__init__(message)
        self.code = code
        self.details = details or {}


HEADER_ALIASES = {
    "model": {"model", "工作内容", "阶段", "stage", "group", "分组"},
    "description": {"description", "事项", "任务", "task", "task name", "任务名称"},
    "kivisense": {"kivisense", "弥知科技", "我方"},
    "brands": {"brands", "brand", "品牌方", "客户", "客户方"},
    "status": {"status", "状态"},
}
DEFAULT_FILL_COLORS = {"FFFFFF", "000000", "D9D9D9", "EDEDED", "F2F2F2", "F5F5F5"}
MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
    7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December",
}


def get_merged_value(ws, row: int, col: int):
    value = ws.cell(row, col).value
    if value is not None:
        return value
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return ws.cell(merged.min_row, merged.min_col).value
    return None


def expand_merged_headers(ws, row: int, start_col: int, end_col: int) -> dict[int, object]:
    """Return one header value per column, expanding horizontal merged ranges."""
    values = {col: ws.cell(row, col).value for col in range(start_col, end_col + 1)}
    for merged in ws.merged_cells.ranges:
        if not (merged.min_row <= row <= merged.max_row):
            continue
        anchor_value = ws.cell(merged.min_row, merged.min_col).value
        for col in range(max(start_col, merged.min_col), min(end_col, merged.max_col) + 1):
            if values[col] is None:
                values[col] = anchor_value
    return values


def _day_number(value) -> int | None:
    """Accept literal day labels only; never treat an Excel serial as a date."""
    if isinstance(value, bool) or isinstance(value, (date, datetime)):
        return None
    try:
        if isinstance(value, str) and not re.fullmatch(r"\s*\d{1,2}\s*", value):
            return None
        number = int(value)
        if float(value) != number:
            return None
    except (TypeError, ValueError):
        return None
    return number if 1 <= number <= 31 else None


def _longest_contiguous_run(columns: list[int]) -> list[int]:
    if not columns:
        return []
    runs: list[list[int]] = []
    current = [columns[0]]
    for col in columns[1:]:
        if col == current[-1] + 1:
            current.append(col)
        else:
            runs.append(current)
            current = [col]
    runs.append(current)
    return max(runs, key=len)


def _header_key(value) -> str | None:
    text = cell_text(value).lower()
    for key, aliases in HEADER_ALIASES.items():
        if text in aliases:
            return key
    return None


def _find_task_name_column(ws, start_row: int, end_row: int, before_col: int) -> int | None:
    candidates: list[tuple[int, int]] = []
    for col in range(1, before_col):
        count = sum(bool(cell_text(ws.cell(row, col).value)) for row in range(start_row, end_row + 1))
        if count:
            candidates.append((count, col))
    return max(candidates)[1] if candidates else None


def detect_header_rows(ws) -> dict:
    """Locate the day-number row in the first 20 rows of a visual Gantt sheet."""
    candidates: list[tuple[int, int, list[int]]] = []
    for row in range(1, min(ws.max_row, 20) + 1):
        day_columns = [col for col in range(1, ws.max_column + 1) if _day_number(ws.cell(row, col).value)]
        run = _longest_contiguous_run(day_columns)
        if len(run) >= 4:
            candidates.append((len(run), row, run))
    if not candidates:
        raise ScheduleImportError("DATE_ROW_NOT_FOUND", "找不到包含日期数字的甘特图时间轴。")
    longest = max(candidate[0] for candidate in candidates)
    best_candidates = [candidate for candidate in candidates if candidate[0] == longest]
    if len(best_candidates) > 1:
        raise ScheduleImportError(
            "AMBIGUOUS_TIMELINE",
            "检测到多个可能的日期时间轴，无法安全导入。",
            {"rows": [candidate[1] for candidate in best_candidates]},
        )
    _, date_row, date_columns = best_candidates[0]
    return {
        "date_row": date_row,
        "month_row": date_row - 1 if date_row > 1 else None,
        "field_header_row": date_row - 1 if date_row > 1 else None,
        "date_columns": date_columns,
    }


def detect_gantt_layout(ws) -> dict:
    """Detect a table-plus-timeline layout without binding it to fixed coordinates."""
    headers = detect_header_rows(ws)
    date_start_col = headers["date_columns"][0]
    date_end_col = headers["date_columns"][-1]
    header_map: dict[str, int] = {}
    header_row = headers["field_header_row"]
    if header_row:
        for col in range(1, date_start_col):
            key = _header_key(get_merged_value(ws, header_row, col))
            if key and key not in header_map:
                header_map[key] = col
    if "description" not in header_map:
        fallback = _find_task_name_column(ws, headers["date_row"] + 1, ws.max_row, date_start_col)
        if fallback:
            header_map["description"] = fallback
    if "description" not in header_map:
        raise ScheduleImportError("TASK_COLUMN_NOT_FOUND", "找不到任务名称列，无法识别甘特图任务。")
    return {
        **headers,
        "date_start_col": date_start_col,
        "date_end_col": date_end_col,
        "task_start_row": headers["date_row"] + 1,
        "header_map": header_map,
    }


def _increment_month(month: int) -> int:
    return 1 if month == 12 else month + 1


def _decrement_month(month: int) -> int:
    return 12 if month == 1 else month - 1


def build_timeline_date_map(ws, layout: dict, import_year: int) -> dict:
    """Map physical timeline columns to date-only values without adding skipped days."""
    columns = list(range(layout["date_start_col"], layout["date_end_col"] + 1))
    date_row = layout["date_row"]
    days = {col: _day_number(ws.cell(date_row, col).value) for col in columns}
    if any(days[col] is None for col in columns):
        raise ScheduleImportError("TIMELINE_GAP", "日期时间轴列必须连续且每列都包含 1 到 31 的日期数字。")

    month_headers = expand_merged_headers(ws, layout["month_row"], columns[0], columns[-1])
    months = {col: month_from_label(month_headers[col]) for col in columns}
    explicit_months = {col for col, month in months.items() if month is not None}
    if not explicit_months:
        raise ScheduleImportError("MONTH_NOT_FOUND", "找不到月份标题，无法建立日期时间轴。")

    # Fill forward from an explicit month. A lower day number means the timeline crossed a month boundary.
    previous_col = None
    for col in columns:
        if months[col] is None and previous_col is not None:
            months[col] = months[previous_col]
            if days[col] < days[previous_col]:
                months[col] = _increment_month(months[col])
        if months[col] is not None:
            previous_col = col

    # A leading blank merged header (July in the supplied workbook) is inferred from the next known month.
    first_known_index = next(index for index, col in enumerate(columns) if months[col] is not None)
    for index in range(first_known_index - 1, -1, -1):
        col = columns[index]
        next_col = columns[index + 1]
        months[col] = months[next_col]
        if days[col] > days[next_col]:
            months[col] = _decrement_month(months[col])

    if any(months[col] is None for col in columns):
        raise ScheduleImportError("MONTH_NOT_FOUND", "月份标题不完整，且无法根据相邻日期推断。")

    dates_by_column: dict[int, date] = {}
    previous_date: date | None = None
    previous_month: int | None = None
    year_offset = 0
    for col in columns:
        month = months[col]
        if previous_month is not None and month < previous_month:
            year_offset += 1
        try:
            current_date = date(import_year + year_offset, month, days[col])
        except ValueError as exc:
            raise ScheduleImportError(
                "INVALID_DAY", f"{get_column_letter(col)} 列的日期无效。", {"column": get_column_letter(col), "day": days[col], "month": month}
            ) from exc
        if previous_date is not None and current_date <= previous_date:
            raise ScheduleImportError(
                "TIMELINE_NOT_INCREASING",
                "日期时间轴必须从左到右严格递增。",
                {"column": get_column_letter(col), "date": current_date.isoformat(), "previous_date": previous_date.isoformat()},
            )
        dates_by_column[col] = current_date
        previous_date = current_date
        previous_month = month

    inferred_months = sorted({months[col] for col in columns if col not in explicit_months})
    return {
        "dates_by_column": dates_by_column,
        "column_date_map": {get_column_letter(col): value.isoformat() for col, value in dates_by_column.items()},
        "inferred_months": [MONTH_NAMES[month] for month in inferred_months],
        "months": [MONTH_NAMES[month] for month in sorted(set(months.values()))],
    }


def _normalize_rgb(value) -> str | None:
    text = str(value or "").upper().lstrip("#")
    if not re.fullmatch(r"[0-9A-F]{6,8}", text):
        return None
    if len(text) == 8:
        text = text[-6:]
    return None if text in DEFAULT_FILL_COLORS else text


def _apply_tint(rgb: str, tint: float) -> str:
    channels = [int(rgb[index : index + 2], 16) for index in range(0, 6, 2)]
    if tint < 0:
        channels = [round(channel * (1 + tint)) for channel in channels]
    else:
        channels = [round(channel + (255 - channel) * tint) for channel in channels]
    return "".join(f"{max(0, min(255, channel)):02X}" for channel in channels)


def _resolve_theme_color(cell, theme: int, tint: float) -> str | None:
    """Resolve an OOXML theme colour to RGB when the workbook carries its theme XML."""
    try:
        workbook = cell.parent.parent
        theme_xml = workbook.loaded_theme
        root = ElementTree.fromstring(theme_xml)
        namespace = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
        scheme = root.find(f".//{namespace}clrScheme")
        if scheme is None or not 0 <= theme < len(scheme):
            return None
        color_node = list(scheme)[theme][0]
        rgb = color_node.attrib.get("lastClr") or color_node.attrib.get("val")
        if not rgb or not re.fullmatch(r"[0-9A-Fa-f]{6}", rgb):
            return None
        return _normalize_rgb(_apply_tint(rgb.upper(), tint))
    except (AttributeError, ElementTree.ParseError, IndexError, TypeError, ValueError):
        return None


def normalize_cell_fill(cell) -> str | None:
    """Normalize RGB, ARGB, theme and indexed solid fills; ignore default backgrounds."""
    fill = cell.fill
    if not fill or (fill.fill_type or getattr(fill, "patternType", None)) != "solid":
        return None
    color = fill.fgColor
    color_type = getattr(color, "type", None)
    if color_type == "rgb":
        return _normalize_rgb(color.rgb)
    if color_type == "indexed":
        index = getattr(color, "indexed", None)
        if not isinstance(index, int) or index in {64, 65} or not 0 <= index < len(COLOR_INDEX):
            return None
        return _normalize_rgb(COLOR_INDEX[index])
    if color_type == "theme":
        theme = getattr(color, "theme", None)
        if theme in {None, 0, 1}:
            return None
        tint = float(getattr(color, "tint", 0) or 0)
        return _resolve_theme_color(cell, theme, tint) or f"theme:{theme}:{tint}"
    return None


def is_timeline_fill(cell) -> bool:
    return normalize_cell_fill(cell) is not None


def normalize_import_status(value) -> str:
    text = cell_text(value).lower()
    if any(token in text for token in ("done", "complete", "completed", "完成", "已完成", "√", "✓")):
        return "done"
    return "incomplete"


def _explicit_year_from_sheet(ws) -> int | None:
    for row in ws.iter_rows():
        for cell in row:
            match = re.search(r"(?<!\d)(20\d{2})(?!\d)", cell_text(cell.value))
            if match:
                return int(match.group(1))
    return None


def _year_from_filename(filename: str) -> int | None:
    match = re.search(r"(?<!\d)(20\d{2})(?!\d)", filename or "")
    return int(match.group(1)) if match else None


def _validated_import_year(value) -> int | None:
    if value in (None, ""):
        return None
    try:
        year = int(value)
    except (TypeError, ValueError) as exc:
        raise ScheduleImportError("INVALID_IMPORT_YEAR", "导入年份必须是四位数字。") from exc
    if not 1900 <= year <= 9999:
        raise ScheduleImportError("INVALID_IMPORT_YEAR", "导入年份必须介于 1900 和 9999 之间。")
    return year


def resolve_import_year(wb, ws, import_year: int | str | None, filename: str = "") -> dict:
    selected_year = _validated_import_year(import_year)
    if selected_year:
        return {"year": selected_year, "year_inferred": False, "source": "user-input", "requires_confirmation": False}
    sheet_year = _explicit_year_from_sheet(ws)
    if sheet_year:
        return {"year": sheet_year, "year_inferred": False, "source": "worksheet", "requires_confirmation": False}
    filename_year = _year_from_filename(filename)
    if filename_year:
        return {"year": filename_year, "year_inferred": True, "source": "filename", "requires_confirmation": True}
    for attribute, source in (("created", "workbook-created-time"), ("modified", "workbook-modified-time")):
        value = getattr(wb.properties, attribute, None)
        if isinstance(value, (date, datetime)) and 1900 <= value.year <= 9999:
            return {"year": value.year, "year_inferred": True, "source": source, "requires_confirmation": True}
    raise ScheduleImportError("YEAR_REQUIRED", "源文件未明确包含年份，请在导入前填写年份后重试。")


def extract_task_segments(ws, row: int, dates_by_column: dict[int, date]) -> dict:
    """Read colored Gantt cells as contiguous physical-column segments."""
    filled_columns = [col for col in sorted(dates_by_column) if normalize_cell_fill(ws.cell(row, col))]
    segments: list[dict] = []
    if filled_columns:
        run_start = filled_columns[0]
        previous_col = filled_columns[0]
        for col in filled_columns[1:] + [None]:
            if col is not None and col == previous_col + 1:
                previous_col = col
                continue
            start = dates_by_column[run_start]
            end = dates_by_column[previous_col]
            segments.append({
                "startDate": start.isoformat(),
                "endDate": end.isoformat(),
                "workingDays": previous_col - run_start + 1,
                "startColumn": get_column_letter(run_start),
                "endColumn": get_column_letter(previous_col),
            })
            if col is not None:
                run_start = col
                previous_col = col
    return {"filled_columns": filled_columns, "segments": segments}


def _project_name_from_sheet(ws) -> str:
    for row in range(1, min(ws.max_row, 2) + 1):
        for col in range(1, min(ws.max_column, 8) + 1):
            value = cell_text(get_merged_value(ws, row, col))
            if value:
                return value
    return ws.title


def parse_gantt_worksheet(ws, import_year: int) -> dict:
    """Parse a visual Gantt worksheet into the existing task model plus import metadata."""
    layout = detect_gantt_layout(ws)
    timeline = build_timeline_date_map(ws, layout, import_year)
    dates_by_column = timeline["dates_by_column"]
    header_map = layout["header_map"]
    tasks: list[dict] = []
    warnings: list[str] = []
    current_model = ""
    model_col = header_map.get("model")
    description_col = header_map["description"]
    kivisense_col = header_map.get("kivisense")
    brands_col = header_map.get("brands")
    status_col = header_map.get("status")

    for row in range(layout["task_start_row"], ws.max_row + 1):
        name = cell_text(ws.cell(row, description_col).value)
        if not name:
            continue
        if model_col:
            model_value = cell_text(get_merged_value(ws, row, model_col))
            if model_value:
                current_model = model_value
        owners = []
        kivisense_value = cell_text(ws.cell(row, kivisense_col).value) if kivisense_col else ""
        brands_value = cell_text(ws.cell(row, brands_col).value) if brands_col else ""
        if kivisense_value:
            owners.append("Kivisense")
        if brands_value:
            owners.append("Brands")
        if not owners:
            owners.append("Kivisense")

        extracted = extract_task_segments(ws, row, dates_by_column)
        filled_columns = extracted["filled_columns"]
        task_warnings: list[str] = []
        if not filled_columns:
            task_warnings.append("任务行没有可识别的甘特图背景色。")
            warnings.append(f"第 {row} 行“{name}”没有可识别的日期。")
            start = end = ""
            duration_calendar_days = 0
        else:
            start_date = dates_by_column[filled_columns[0]]
            end_date = dates_by_column[filled_columns[-1]]
            start = start_date.isoformat()
            end = end_date.isoformat()
            duration_calendar_days = (end_date - start_date).days + 1
        source = {
            "sheet": ws.title,
            "row": row,
            "startColumn": get_column_letter(filled_columns[0]) if filled_columns else None,
            "endColumn": get_column_letter(filled_columns[-1]) if filled_columns else None,
        }
        tasks.append({
            "model": current_model,
            "group": current_model,
            "name": name,
            "taskName": name,
            "kivisense": kivisense_value or None,
            "brands": brands_value or None,
            "owners": owners,
            "status": normalize_import_status(ws.cell(row, status_col).value) if status_col else "incomplete",
            "start": start,
            "end": end,
            "startDate": start,
            "endDate": end,
            "durationWorkingDays": len(filled_columns),
            "durationCalendarDays": duration_calendar_days,
            "segments": extracted["segments"],
            "source": source,
            "warnings": task_warnings,
        })

    if not tasks:
        raise ScheduleImportError("TASKS_NOT_FOUND", "未在甘特图中识别到任何任务名称。")
    return {
        "project_name": _project_name_from_sheet(ws),
        "tasks": tasks,
        "include_model": bool(model_col),
        "include_status": bool(status_col),
        "layout": layout,
        "timeline": timeline,
        "warnings": warnings,
    }


def validate_imported_schedule(parsed: dict) -> dict:
    """Validate date-only task data and return a compact preview summary."""
    valid_tasks = []
    for task in parsed["tasks"]:
        start = task.get("start", "")
        end = task.get("end", "")
        if not start or not end:
            continue
        try:
            start_date = date.fromisoformat(start)
            end_date = date.fromisoformat(end)
        except ValueError as exc:
            raise ScheduleImportError("INVALID_TASK_DATE", f"任务“{task['name']}”包含无效日期。") from exc
        if end_date < start_date:
            raise ScheduleImportError("INVALID_TASK_RANGE", f"任务“{task['name']}”的结束日期早于开始日期。")
        valid_tasks.append(task)
    if not valid_tasks:
        raise ScheduleImportError("TASK_DATES_NOT_FOUND", "没有识别到带有效日期的任务，请检查甘特图填充色和日期轴。")
    return {
        "task_count": len(parsed["tasks"]),
        "valid_date_task_count": len(valid_tasks),
        "task_date_range": f"{min(task['start'] for task in valid_tasks)} 至 {max(task['end'] for task in valid_tasks)}",
    }


def parse_imported_workbook(file_bytes: bytes, filename: str = "", import_year: int | str | None = None) -> dict:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    errors: list[ScheduleImportError] = []
    for ws in wb.worksheets:
        try:
            year_resolution = resolve_import_year(wb, ws, import_year, filename)
            parsed = parse_gantt_worksheet(ws, year_resolution["year"])
            validation = validate_imported_schedule(parsed)
        except ScheduleImportError as exc:
            errors.append(exc)
            continue

        warnings = list(parsed["warnings"])
        if parsed["timeline"]["inferred_months"]:
            warnings.append(f"已根据相邻月份和日期回绕推断：{', '.join(parsed['timeline']['inferred_months'])}。")
        if year_resolution["year_inferred"]:
            warnings.append(f"源文件未明确存储年份，当前按 {year_resolution['year']} 年解析，请确认。")
        column_dates = list(parsed["timeline"]["column_date_map"].values())
        return {
            "project_name": parsed["project_name"],
            "tasks": parsed["tasks"],
            "include_model": parsed["include_model"],
            "include_status": parsed["include_status"],
            "yearInferred": year_resolution["year_inferred"],
            "inferredYearSource": year_resolution["source"],
            "requires_year_confirmation": year_resolution["requires_confirmation"],
            "warnings": warnings,
            "import_preview": {
                "sheet": ws.title,
                "year": year_resolution["year"],
                "year_inferred": year_resolution["year_inferred"],
                "inferred_year_source": year_resolution["source"],
                "months": parsed["timeline"]["months"],
                "inferred_months": parsed["timeline"]["inferred_months"],
                "column_date_map": parsed["timeline"]["column_date_map"],
                "date_range": f"{column_dates[0]} 至 {column_dates[-1]}",
                "task_date_range": validation["task_date_range"],
                "task_count": validation["task_count"],
                "valid_date_task_count": validation["valid_date_task_count"],
                "warnings": warnings,
            },
        }
    if errors:
        raise errors[0]
    raise ScheduleImportError("WORKSHEET_NOT_FOUND", "工作簿中没有可识别的排期工作表。")


def parse_raw_tasks(raw_text: str, include_model: bool = False) -> list[dict]:
    tasks: list[dict] = []
    for line_number, raw_line in enumerate(raw_text.splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        line = re.sub(r"^\s*\d+[\.\)、)]\s*", "", line)
        dates = DATE_RE.findall(line)
        duration_match = DURATION_RE.search(line)
        if not dates:
            raise ValueError(f"第 {line_number} 行缺少开始日期")

        start = dates[0]
        end = dates[1] if len(dates) > 1 else ""
        workdays = re.search(r"\d+", duration_match.group(0)).group(0) if duration_match else ""
        owners = []
        if re.search(r"\b(kivisense|kv)\b|我方", line, re.IGNORECASE):
            owners.append("Kivisense")
        if re.search(r"\b(brand|brands|client)\b|客户|品牌方", line, re.IGNORECASE):
            owners.append("Brands")

        first_date = DATE_RE.search(line)
        name_part = line[: first_date.start()]
        pieces = [piece.strip() for piece in re.split(r"[,，/、]+", name_part) if piece.strip()]
        filtered = [piece for piece in pieces if not OWNER_RE.fullmatch(piece)]
        if include_model:
            if len(filtered) < 2:
                raise ValueError(f"第 {line_number} 行开启 Model 后，需要写：Model, 事项名称, 相关方, 开始日期, 结束日期或工作日")
            model = filtered[0]
            name = filtered[1]
        else:
            model = ""
            name = filtered[0] if filtered else pieces[0] if pieces else ""
        if not name:
            raise ValueError(f"第 {line_number} 行缺少事项名称")

        task = {"model": model, "name": name, "owners": owners, "start": start}
        if end:
            task["end"] = end
        if workdays:
            task["workdays"] = int(workdays)
        tasks.append(task)

    if not tasks:
        raise ValueError("请至少输入一条事项")
    return tasks


class TimelineRequestHandler(BaseHTTPRequestHandler):
    server_version = "TimelineMaker/1.0"

    def do_GET(self) -> None:
        self.handle_get_like_request(write_body=True)

    def do_HEAD(self) -> None:
        self.handle_get_like_request(write_body=False)

    def handle_get_like_request(self, write_body: bool) -> None:
        parsed = urlparse(self.path)
        if BASE_PATH and parsed.path == BASE_PATH:
            location = with_base("/") + (f"?{parsed.query}" if parsed.query else "")
            self.respond_redirect(location)
            return

        path = strip_base_path(parsed.path)
        if path == "/":
            self.respond_text(render_index_html(), "text/html; charset=utf-8", write_body=write_body)
            return
        if path == "/assets/kivisense-logo.png":
            self.respond_file(Path(__file__).parent / "assets" / "kivisense-logo.png", "image/png", write_body=write_body)
            return
        self.send_error(404)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        path = strip_base_path(parsed.path)
        try:
            length = int(self.headers.get("content-length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))

            if path == "/import":
                file_b64 = payload.get("file_b64") or ""
                if not file_b64:
                    raise ValueError("请先选择一个 .xlsx 排期文件。")
                file_bytes = base64.b64decode(file_b64)
                imported = parse_imported_workbook(
                    file_bytes,
                    filename=payload.get("filename") or "",
                    import_year=payload.get("import_year"),
                )
                self.respond_json(imported)
                return

            if path != "/generate":
                self.send_error(404)
                return

            include_model = True
            include_status = bool(payload.get("include_status", True))
            structured_tasks = payload.get("tasks")
            tasks = structured_tasks if structured_tasks else parse_raw_tasks(payload.get("raw_tasks", ""), include_model=include_model)
            config = {
                "project_name": payload.get("project_name") or "Timeline",
                "tasks": tasks,
                "include_model": include_model,
                "include_status": include_status,
                "language": payload.get("language") or "zh",
            }
            workbook = build_workbook(config)
            output = BytesIO()
            workbook.save(output)
            body = output.getvalue()
            filename = f"timeline_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except ScheduleImportError as exc:
            self.respond_json({"error": str(exc), "code": exc.code, "details": exc.details}, status=400)
        except Exception as exc:
            self.respond_json({"error": str(exc)}, status=400)

    def respond_text(self, text: str, content_type: str, write_body: bool = True) -> None:
        body = text.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        if write_body:
            self.wfile.write(body)

    def respond_redirect(self, location: str, status: int = 301) -> None:
        self.send_response(status)
        self.send_header("Location", location)
        self.end_headers()

    def respond_json(self, payload: dict, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def respond_file(self, path: Path, content_type: str, write_body: bool = True) -> None:
        if not path.exists():
            self.send_error(404)
            return
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        if write_body:
            self.wfile.write(body)

    def log_message(self, format: str, *args) -> None:
        return


def run_self_test() -> Path:
    config = {
        "project_name": "Local Test",
        "include_model": True,
        "include_status": False,
        "tasks": [
            {"model": "需求", "name": "Project requirement", "owners": ["Kivisense"], "start": "2026-06-01", "workdays": 5},
            {"model": "设计", "name": "Creative Proposal", "owners": ["Kivisense", "Brands"], "start": "2026-06-08", "end": "2026-06-22"},
            {"model": "需求", "name": "Scope addendum", "owners": ["Brands"], "start": "2026-06-18", "workdays": 4},
        ],
    }
    workbook = build_workbook(config)
    output_path = Path(tempfile.gettempdir()) / "timeline_maker_self_test.xlsx"
    workbook.save(output_path)
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Run Timeline Maker locally.")
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument("--no-browser", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()

    if args.self_test:
        output_path = run_self_test()
        print(output_path)
        return 0

    server = ThreadingHTTPServer((HOST, args.port), TimelineRequestHandler)
    url = f"http://{HOST}:{args.port}{with_base('/')}"
    print(f"Timeline Maker is running: {url}")
    if BASE_PATH:
        print(f"BASE_PATH is set to: {BASE_PATH}")
    if not args.no_browser:
        webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
