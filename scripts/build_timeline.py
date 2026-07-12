#!/usr/bin/env python3
"""Build a Kivisense-style timeline workbook from JSON input."""

from __future__ import annotations

import json
import re
import sys
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU


APPROVED_PALETTE = ["A9D08E", "F8CBAD", "7030A0", "DC888E", "757171", "5B9BD5", "00B050"]
FORBIDDEN_YELLOWS = {"FFC000", "FFFF00", "FFFFFF00", "FFFF99", "FFF2CC"}
TIMELINE_CELL_PX = 30
TIMELINE_COLUMN_WIDTH = 3.6
TIMELINE_ROW_HEIGHT_PT = 22.5
EXPORT_FONT_NAME = "Gotham"
CATEGORY_COLORS = {
    "planning": ["A9D08E", "5B9BD5", "F8CBAD"],
    "compliance": ["F8CBAD", "757171", "5B9BD5"],
    "development": ["7030A0", "5B9BD5", "757171"],
    "design": ["DC888E", "757171", "F8CBAD"],
    "special": ["7030A0", "DC888E", "5B9BD5"],
    "testing": ["A9D08E", "5B9BD5", "DC888E"],
    "launch": ["00B050", "7030A0", "5B9BD5"],
    "default": ["A9D08E", "F8CBAD", "5B9BD5"],
}
CHINA_PUBLIC_HOLIDAYS = {
    # 2026 China public holiday and adjusted-workday calendar.
    "2026-01-01", "2026-01-02", "2026-01-03",
    "2026-02-15", "2026-02-16", "2026-02-17", "2026-02-18", "2026-02-19", "2026-02-20", "2026-02-21", "2026-02-22", "2026-02-23",
    "2026-04-04", "2026-04-05", "2026-04-06",
    "2026-05-01", "2026-05-02", "2026-05-03", "2026-05-04", "2026-05-05",
    "2026-06-19", "2026-06-20", "2026-06-21",
    "2026-09-25", "2026-09-26", "2026-09-27",
    "2026-10-01", "2026-10-02", "2026-10-03", "2026-10-04", "2026-10-05", "2026-10-06", "2026-10-07",
}
CHINA_ADJUSTED_WORKDAYS = {
    "2026-01-04",
    "2026-02-14", "2026-02-28",
    "2026-05-09",
    "2026-09-20", "2026-10-10",
}


def parse_date(value: str) -> date:
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unsupported date format: {value!r}")


def date_key(day: date) -> str:
    return day.isoformat()


def is_workday(day: date) -> bool:
    key = date_key(day)
    if key in CHINA_ADJUSTED_WORKDAYS:
        return True
    if key in CHINA_PUBLIC_HOLIDAYS:
        return False
    return day.weekday() < 5


def next_workday(day: date) -> date:
    while not is_workday(day):
        day += timedelta(days=1)
    return day


def add_workdays(start: date, workdays: int) -> date:
    if workdays < 1:
        raise ValueError("workdays must be >= 1")
    current = next_workday(start)
    remaining = workdays - 1
    while remaining:
        current += timedelta(days=1)
        if is_workday(current):
            remaining -= 1
    return current


def workday_range(start: date, end: date) -> list[date]:
    days = []
    current = start
    while current <= end:
        if is_workday(current):
            days.append(current)
        current += timedelta(days=1)
    return days


def count_workdays(start: date, end: date) -> int:
    return len(workday_range(start, end))


def normalize_owners(owners) -> set[str]:
    if owners is None:
        return set()
    if isinstance(owners, str):
        parts = re.split(r"[,，/+&、\s]+", owners)
    else:
        parts = [str(item) for item in owners]
    result = set()
    for part in parts:
        token = part.strip().lower()
        if token in {"kivisense", "kv", "我方"}:
            result.add("Kivisense")
        if token in {"brand", "brands", "client", "客户", "品牌方"}:
            result.add("Brands")
    return result


def infer_category(task: dict) -> str:
    explicit = str(task.get("category", "")).strip().lower()
    if explicit in CATEGORY_COLORS:
        return explicit
    name = str(task.get("name", "")).lower()
    if any(k in name for k in ("privacy", "compliance", "legal")):
        return "compliance"
    if any(k in name for k in ("development", "integration", "frontend", "backend", "开发", "联调")):
        return "development"
    if any(k in name for k in ("design", "creative", "content", "video", "mural", "设计", "素材")):
        return "design"
    if any(k in name for k in ("easter", "coupon", "middle office")):
        return "special"
    if any(k in name for k in ("uat", "test", "testing", "setup", "测试", "验收")):
        return "testing"
    if any(k in name for k in ("launch", "online", "go live", "上线")):
        return "launch"
    return "default"


def normalize_color(value: str | None) -> str | None:
    if not value:
        return None
    color = str(value).strip().lstrip("#").upper()
    if len(color) == 8 and color.startswith("FF"):
        color = color[2:]
    if len(color) != 6 or not re.fullmatch(r"[0-9A-F]{6}", color):
        return None
    if color in FORBIDDEN_YELLOWS:
        return None
    return color


def assign_bar_colors(tasks: list[dict]) -> None:
    recent: list[str] = []
    for task in tasks:
        explicit = normalize_color(task.get("color"))
        category = task.get("category", "default")
        candidates = []
        if explicit:
            candidates.append(explicit)
        candidates.extend(CATEGORY_COLORS.get(category, CATEGORY_COLORS["default"]))
        candidates.extend(APPROVED_PALETTE)
        chosen = next((color for color in candidates if color not in recent[-3:] and color not in FORBIDDEN_YELLOWS), None)
        task["bar_color"] = chosen or APPROVED_PALETTE[0]
        recent.append(task["bar_color"])


def group_tasks_by_model(tasks: list[dict], default_model: str) -> list[dict]:
    model_order: dict[str, int] = {}
    for task in tasks:
        model = task.get("model") or default_model
        task["model"] = model
        if model not in model_order:
            model_order[model] = len(model_order)
    return sorted(tasks, key=lambda task: (model_order[task["model"]], task["start"], task["original_index"]))


def normalize_tasks(raw_tasks: list[dict]) -> list[dict]:
    tasks = []
    for index, raw in enumerate(raw_tasks, start=1):
        name = str(raw.get("name") or raw.get("task") or raw.get("事项名称") or "").strip()
        if not name:
            raise ValueError(f"Task #{index} is missing a name")
        start = parse_date(raw.get("start") or raw.get("start_date") or raw.get("开始日期"))
        start = next_workday(start)
        end_value = raw.get("end") or raw.get("end_date") or raw.get("结束日期")
        workdays_value = raw.get("workdays") or raw.get("duration") or raw.get("天数") or raw.get("工作日")
        workdays_match = re.search(r"\d+", str(workdays_value)) if workdays_value is not None else None
        if end_value:
            end = next_workday(parse_date(end_value))
            if end < start:
                raise ValueError(f"Task #{index} end date is before start date")
            workdays = count_workdays(start, end)
        elif workdays_match:
            workdays = int(workdays_match.group(0))
            end = add_workdays(start, workdays)
        else:
            raise ValueError(f"Task #{index} is missing end date or workday duration")
        owners = normalize_owners(raw.get("owners") or raw.get("owner") or raw.get("责任人"))
        status = str(raw.get("status", "incomplete")).strip().lower()
        tasks.append(
            {
                "name": name,
                "owners": owners,
                "start": start,
                "end": end,
                "workdays": workdays,
                "status": status,
                "category": infer_category(raw),
                "color": raw.get("color") or raw.get("颜色"),
                "model": str(raw.get("model") or raw.get("stage") or raw.get("module") or raw.get("工作内容") or raw.get("阶段") or "").strip(),
                "original_index": index,
            }
        )
    return tasks


def month_label(day: date, language: str = 'en') -> str:
    if language == 'zh':
        return f'{day.month}月'
    return day.strftime("%B")


def one_cell_image_anchor(col: int, row: int, width_px: int, height_px: int, x_offset_px: float = 0, y_offset_px: float = 0) -> OneCellAnchor:
    marker = AnchorMarker(
        col=col - 1,
        colOff=pixels_to_EMU(x_offset_px),
        row=row - 1,
        rowOff=pixels_to_EMU(y_offset_px),
    )
    return OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(pixels_to_EMU(width_px), pixels_to_EMU(height_px)),
    )


def build_workbook(config: dict) -> Workbook:
    language = str(config.get("language") or "en").lower()
    if language.startswith("zh") or language in {"cn", "chinese", "中文"}:
        language = "zh"
    else:
        language = "en"
    tasks = normalize_tasks(config.get("tasks", []))
    if not tasks:
        raise ValueError("Input JSON must include at least one task")
    include_model = True
    include_status = bool(config.get("include_status", True))
    if include_model:
        tasks = group_tasks_by_model(tasks, "未分类" if language == "zh" else "Uncategorized")
    assign_bar_colors(tasks)

    min_start = min(task["start"] for task in tasks)
    max_end = max(task["end"] for task in tasks)
    buffer_end = add_workdays(max_end + timedelta(days=1), int(config.get("buffer_workdays", 5)))
    days = workday_range(min_start, buffer_end)

    wb = Workbook()
    ws = wb.active
    ws.title = "Timeline"
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", start_color="FF000000", end_color="FF000000")
    title_font = Font(name=EXPORT_FONT_NAME, size=16)
    body_font = Font(name=EXPORT_FONT_NAME, size=11)
    model_font = Font(name=EXPORT_FONT_NAME, size=11, scheme=None)
    header_font = Font(name=EXPORT_FONT_NAME, size=11, color="FFFFFFFF", bold=True)

    labels = {
        "zh": {"model": "工作内容", "description": "事项", "kivisense": "弥知科技", "brands": "品牌方", "status": "状态", "done": "完成", "incomplete": "未完成"},
        "en": {"model": "Model", "description": "Description", "kivisense": "Kivisense", "brands": "Brands", "status": "Status", "done": "Done", "incomplete": "Incomplete"},
    }[language]

    project_name = config.get("project_name", "AR Campaign")
    literal_incomplete = bool(config.get("literal_incomplete_status", False))

    header_defs = []
    if include_model:
        header_defs.append(("model", labels["model"]))
    header_defs.extend([("description", labels["description"]), ("kivisense", labels["kivisense"]), ("brands", labels["brands"])])
    if include_status:
        header_defs.append(("status", labels["status"]))
    column_index = {key: index for index, (key, _label) in enumerate(header_defs, start=1)}
    date_start_col = len(header_defs) + 1
    ws.freeze_panes = f"{get_column_letter(date_start_col)}5"

    last_col = len(header_defs) + len(days)
    last_row = 4 + len(tasks)

    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=4)
    ws.cell(1, 2).value = str(project_name)[:20]
    ws.cell(1, 2).font = title_font
    ws.cell(1, 2).alignment = Alignment(horizontal="center", vertical="center")

    logo_path = Path(config.get("logo_path") or Path(__file__).resolve().parents[1] / "assets" / "kivisense-logo.png")
    star_path = Path(config.get("star_path") or Path(__file__).resolve().parents[1] / "assets" / "gantt-end-star.png")
    if logo_path.exists():
        try:
            logo = Image(str(logo_path))
            logo_width = 235
            logo_height = 98
            logo.width = logo_width
            logo.height = logo_height
            logo.anchor = one_cell_image_anchor(1, 1, logo_width, logo_height, 20, 0)
            ws.add_image(logo)
        except Exception:
            # If Pillow is not installed or cannot load image assets, keep export working.
            # Installing requirements.txt will restore embedded logo/image output.
            pass

    for col, (_key, header) in enumerate(header_defs, start=1):
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
        cell = ws.cell(3, col)
        cell.value = header
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.border = border

    month_start_col = date_start_col
    current_month = days[0].month
    for offset, day in enumerate(days, start=date_start_col):
        ws.cell(4, offset).value = day.day
        ws.cell(4, offset).font = header_font
        ws.cell(4, offset).fill = header_fill
        ws.cell(4, offset).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(4, offset).border = border
        if day.month != current_month:
            ws.merge_cells(start_row=3, start_column=month_start_col, end_row=3, end_column=offset - 1)
            ws.cell(3, month_start_col).value = month_label(days[month_start_col - date_start_col], language)
            ws.cell(3, month_start_col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(3, month_start_col).fill = header_fill
            ws.cell(3, month_start_col).font = header_font
            current_month = day.month
            month_start_col = offset
    ws.merge_cells(start_row=3, start_column=month_start_col, end_row=3, end_column=last_col)
    ws.cell(3, month_start_col).value = month_label(days[month_start_col - date_start_col], language)
    ws.cell(3, month_start_col).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(3, month_start_col).fill = header_fill
    ws.cell(3, month_start_col).font = header_font

    day_to_col = {day: index + date_start_col for index, day in enumerate(days)}
    for row_offset, task in enumerate(tasks, start=5):
        if include_model:
            ws.cell(row_offset, column_index["model"]).value = task["model"]

        description_col = column_index["description"]
        ws.cell(row_offset, description_col).value = task["name"]
        ws.cell(row_offset, description_col).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row_offset, description_col).font = body_font

        if "Kivisense" in task["owners"]:
            ws.cell(row_offset, column_index["kivisense"]).value = "√"
        if "Brands" in task["owners"]:
            ws.cell(row_offset, column_index["brands"]).value = "√"

        if include_status:
            completed_tokens = {"complete", "completed", "done", "完成", "已完成", "√"}
            if task["status"] in completed_tokens:
                ws.cell(row_offset, column_index["status"]).value = labels["done"] if language == "zh" else "√"
            elif literal_incomplete:
                ws.cell(row_offset, column_index["status"]).value = labels["incomplete"]

        fill = PatternFill("solid", fgColor=task["bar_color"])
        for day in workday_range(task["start"], task["end"]):
            col = day_to_col.get(day)
            if col:
                ws.cell(row_offset, col).fill = fill

        end_col = day_to_col.get(task["end"])
        if end_col:
            image_added = False
            if star_path.exists():
                try:
                    star = Image(str(star_path))
                    star_width = 25
                    star_height = 27
                    star.width = star_width
                    star.height = star_height
                    star.anchor = one_cell_image_anchor(
                        end_col,
                        row_offset,
                        star_width,
                        star_height,
                        max((TIMELINE_CELL_PX - star_width) / 2, 0),
                        max((TIMELINE_CELL_PX - star_height) / 2, 0),
                    )
                    ws.add_image(star)
                    image_added = True
                except Exception:
                    image_added = False
            if not image_added:
                star_cell = ws.cell(row_offset, end_col)
                star_cell.value = "★"
                star_cell.font = Font(name=EXPORT_FONT_NAME, size=11, bold=True, color="C00000")
                star_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, last_col + 1):
            c = ws.cell(row_offset, col)
            c.border = border
            font = copy(c.font)
            font.name = EXPORT_FONT_NAME
            font.size = 11
            font.scheme = None
            font.family = None
            font.charset = None
            c.font = font
            if col != description_col:
                c.alignment = Alignment(horizontal="center", vertical="center")

    if include_model:
        model_col = column_index["model"]
        group_start = 5
        while group_start <= last_row:
            model = ws.cell(group_start, model_col).value
            group_end = group_start
            while group_end + 1 <= last_row and ws.cell(group_end + 1, model_col).value == model:
                group_end += 1
            if group_end > group_start:
                ws.merge_cells(start_row=group_start, start_column=model_col, end_row=group_end, end_column=model_col)
            # OOXML keeps a merged range's visible style on its top-left anchor cell.
            cell = ws.cell(group_start, model_col)
            cell.value = model
            cell.font = copy(model_font)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            group_start = group_end + 1


    # Force the full two-row header band to render as black background with white text,
    # including cells inside merged ranges across Excel, Numbers and Google Sheets.
    for row in range(3, 5):
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.fill = copy(header_fill)
            cell.font = copy(header_font)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for header, col in column_index.items():
        letter = get_column_letter(col)
        if header == "description":
            ws.column_dimensions[letter].width = 47
        elif header == "model":
            ws.column_dimensions[letter].width = 28
        else:
            ws.column_dimensions[letter].width = 14
    for col in range(date_start_col, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = TIMELINE_COLUMN_WIDTH
    ws.row_dimensions[1].height = 46
    ws.row_dimensions[2].height = 46
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 28
    for row in range(5, last_row + 1):
        ws.row_dimensions[row].height = TIMELINE_ROW_HEIGHT_PT

    return wb


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: build_timeline.py input.json output.xlsx", file=sys.stderr)
        return 2
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    config = json.loads(input_path.read_text(encoding="utf-8"))
    wb = build_workbook(config)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
